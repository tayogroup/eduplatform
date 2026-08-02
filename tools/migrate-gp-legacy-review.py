"""One-off: salvage a review made against the FIRST Global Perspectives export.

The workbook reviewed on 2026-08-01 was taken from the export that shipped in
c0aed960 — before callout ids gained their "box-" prefix and before a cell's
fields were labelled "[Answer]" on their own line. apply-ehel-global-
perspectives-script-review.py reads the current format, so it cannot attribute
edits in that file, and the review represents real work that should not be
redone.

WHAT THIS DOES NOT DO
=====================
Most of that review — 2,524 of 2,851 edits — was one instruction: stop speaking
the on-screen heading. A callout box had its title line deleted; a glossary clip
had its term stripped from "term. meaning". That is a change to what the
narration SAYS, not to what the page shows, so it belongs in course-ui.js and
tools/lib/ehel-global-perspectives-narration.js and is implemented there. Writing
it as 2,524 content overrides would have deleted the headings off the screen,
which is the opposite of what was asked.

What is left is the ~327 genuine text edits: en/em dashes, table cell spacing,
and punctuation added for spoken pauses. Those are content, and this maps them
onto the current JSON paths.

HOW AN EDIT IS ATTRIBUTED
=========================
The old export joined several fields into one cell with no labels, so the cell
alone cannot say which field changed. But the field values are known from the
JSON, so the old cell can be reconstructed block by block and the reviewed cell
aligned against it: a block that still matches is untouched, and the edit is
isolated to the ones that do not. A row that cannot be aligned exactly is
reported and skipped rather than guessed at.

Usage:
    python tools/migrate-gp-legacy-review.py --workbook <reviewed.xlsx>
        --baseline <baseline.xlsx> [--out <path.json>] [--dry]

`--baseline` is the workbook the reviewer actually worked from, re-rendered from
the commit they had. Produce it with:
    git worktree add /tmp/gp-v1 c0aed960
    python /tmp/gp-v1/tools/export-ehel-global-perspectives-scripts.py --out <baseline.xlsx>
"""

from __future__ import annotations

import argparse
import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COURSE = ROOT / "src" / "prototypes" / "ehel-academy" / "global-perspectives"
DEFAULT_OUT = COURSE / "data" / "script-review.json"

_spec = importlib.util.spec_from_file_location(
    "export_gp_scripts", ROOT / "tools" / "export-ehel-global-perspectives-scripts.py")
export = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(export)

# Categories whose whole point was the heading removal, now handled in the
# narration layer. An edit here that is ONLY the heading removal is not a
# content change and is skipped silently; anything more is still salvaged.
BOX_CATEGORIES = {
    "Big idea (narrated)", "Worked example (narrated)", "AI tutor prompt (narrated)",
    "Speaking prompt (narrated)", "Teacher session (narrated)", "Reflection box (narrated)",
    "Activity box (narrated)",
}
GLOSSARY = "Glossary word (narrated)"

# Categories whose old cell did NOT contain the section title: it was written
# only into the "Title / Word" column. Today those rows carry a title field too,
# so it has to be excluded when lining the two up.
TITLE_OUTSIDE_CELL = {
    "Grown-up guide (adult voice)", "Mini-project step", "Skills toolkit",
    "Checklist", "Activity",
}


# Two rows the automatic alignment cannot resolve, with the boundary stated
# outright instead of guessed. Both change structure, not just characters:
#
#   gp-g01-u04-guide-12  the reviewer merged the last two lines of the water
#                        chant into one, so the cell lost a line and no
#                        line-for-line mapping holds.
#   gp-g02-u03-step-7    the intro and the table both changed at once, leaving
#                        no untouched block to anchor on.
#
# Each entry slices the REVIEWED cell by line, so the text applied is the
# reviewer's own — only where one field ends and the next begins is asserted
# here. The table field of step-7 is deliberately absent: that edit removed
# trailing spaces from a rendered row whose underlying cells are empty either
# way, so there is nothing in the data to change.
MANUAL_SPLITS: dict[str, list[tuple[str, int, int | None]]] = {
    "gp-g01-u04-guide-12": [
        ("grownUpGuide.sections.11.body", 0, 7),
        ("grownUpGuide.sections.11.items", 7, None),
    ],
    "gp-g02-u03-step-7": [
        ("project.steps.6.intro", 0, 5),
    ],
}


def rows_for(grade: int, root: Path | None = None) -> dict[str, tuple[list[str], list, dict]]:
    """Rows for a grade, keyed by item id, with the legacy id forms aliased.

    `root` selects which checkout's unit data to read. Decomposing an old cell
    has to use the data that PRODUCED it: once a first pass has been applied and
    the course rebuilt, current content no longer matches the workbook's
    baseline, and every row whose fix already landed stops aligning — reported
    as a failure when it is in fact a success. Passing the baseline checkout
    here keeps the two apart. The PATHS are unaffected either way, because they
    are positional and neither pass changes how many items a unit has.
    """
    out: dict[str, tuple[list[str], list, dict]] = {}
    files = export.unit_files(grade) if root is None else sorted(
        (root / "src" / "prototypes" / "ehel-academy" / "global-perspectives"
         / f"grade-{grade}" / "data" / "units").glob("unit-*.json"),
        key=lambda p: int(p.stem.split("-")[1]),
    )
    for path in files:
        unit = json.loads(path.read_text(encoding="utf-8"))
        for row, layout, source in export.unit_rows_with_sources(unit):
            out[row[2]] = (row, layout, source)
            # The callout ids gained a "box-" prefix after that export; the
            # reviewed workbook still carries the old form.
            if "-box-" in row[2]:
                out[row[2].replace("-box-", "-", 1)] = (row, layout, source)
            # The grown-up guide's notes row was singular in that export.
            if row[2].endswith("-guide-notes"):
                out[row[2][: -len("s")]] = (row, layout, source)
    return out


def parse_v1_labeled(cell: str, layout: list) -> dict[str, str] | None:
    """Read a cell that the OLD export wrote with "Label: " prefixes.

    Not every category was unlabelled back then. Practice, the unit quiz,
    Reflection and the Challenge already carried labels ("Question: …",
    "One way to answer: …"), so treating their cells as a bare concatenation of
    field values never matched and every one of those rows was reported
    unalignable. Where the labels ARE present, both the baseline and the
    reviewed cell parse the same way and the fields can simply be compared —
    no alignment guesswork at all.

    Returns None when the cell does not carry this shape.
    """
    labelled = [(label, path) for label, path, _kind in layout if label]
    if not labelled:
        return None
    lines = cell.split("\n")
    anchors: list[tuple[int, str]] = []
    cursor = 0
    for label, path in labelled:
        prefix = f"{label}: "
        found = next((i for i in range(cursor, len(lines)) if lines[i].startswith(prefix)), None)
        if found is None:
            continue
        anchors.append((found, path))
        cursor = found + 1
    if not anchors:
        return None
    if anchors[0][0] != 0:
        return None  # text before the first label: not this shape
    out: dict[str, str] = {}
    for index, (start, path) in enumerate(anchors):
        stop = anchors[index + 1][0] if index + 1 < len(anchors) else len(lines)
        label = next(l for l, p in labelled if p == path)
        value = "\n".join(lines[start:stop])[len(label) + 2:].strip()
        if value:
            out[path] = value
    return out or None


def align(old_cell: str, new_cell: str, source: dict[str, str], layout: list,
          skip_title: bool = False) -> dict[str, str] | None:
    """Work out which field(s) a reviewed cell changed.

    The old cell was the non-empty field values joined by newlines, in layout
    order. Walk them in order: a field whose text is still present at the front
    of what remains is unchanged; the first that is not absorbs the difference.
    Returns None when the alignment is not exact.
    """
    # Categories the old export already labelled parse exactly on both sides,
    # so compare field by field rather than guessing at block boundaries.
    old_labelled = parse_v1_labeled(old_cell, layout)
    if old_labelled is not None and old_labelled == {k: v for k, v in source.items() if k in old_labelled}:
        new_labelled = parse_v1_labeled(new_cell, layout)
        if new_labelled is not None:
            changed = {p: v for p, v in new_labelled.items() if source.get(p, "") != v}
            if changed:
                return changed

    present = [path for _label, path, _kind in layout if path in source]
    # The old export put a section's title only in the "Title / Word" COLUMN,
    # never in the Script Text cell — the title field came later. Today's layout
    # has it, so leaving it in makes every multi-block row fail to align at its
    # very first field.
    if skip_title:
        # ".label" as well as ".title": an Activity's heading is stored under
        # `activities.N.label`, so filtering only on ".title" left it in and
        # every Activity row failed at its very first field.
        present = [path for path in present if not path.endswith((".title", ".label"))]
    if not present:
        return None
    if len(present) == 1:
        return {present[0]: new_cell.strip()}

    remaining = new_cell
    edits: dict[str, str] = {}
    for index, path in enumerate(present):
        old_value = source[path]
        if remaining.startswith(old_value):
            remaining = remaining[len(old_value):].lstrip("\n")
            continue
        # This field changed. If it is the last one, it takes the rest.
        if index == len(present) - 1:
            edits[path] = remaining.strip()
            return edits if edits else None
        # Otherwise the following fields must still match at the tail, so the
        # changed field is what sits between the front and that tail.
        tail = "\n".join(source[p] for p in present[index + 1:])
        if not remaining.endswith(tail):
            # More than one field changed, so there is no untouched tail to
            # anchor on. Fall through to the line-span method rather than
            # giving up — returning here left 28 rows unsalvaged.
            break
        edits[path] = remaining[: len(remaining) - len(tail)].strip()
        return edits if edits else None
    if edits:
        return edits

    # Prefix/suffix alignment only works when a single field changed. These
    # edits are punctuation — dashes, spacing, an added comma — so they change
    # characters without adding or removing lines. When the line count is
    # unchanged, the old blocks' line spans still hold and each field can be
    # rebuilt from the same span of the reviewed cell. This is what recovers
    # the multi-block rows (the grown-up guide, the mini-project steps).
    old_lines = old_cell.split("\n")
    new_lines = new_cell.split("\n")
    if len(old_lines) != len(new_lines):
        return None
    cursor = 0
    rebuilt: dict[str, str] = {}
    for path in present:
        span = len(source[path].split("\n"))
        if old_lines[cursor:cursor + span] != source[path].split("\n"):
            return None  # the old cell does not decompose the way we think
        chunk = "\n".join(new_lines[cursor:cursor + span]).strip()
        if chunk != source[path]:
            rebuilt[path] = chunk
        cursor += span
        # Blocks were joined by a single newline; skip it.
        if cursor < len(old_lines) and old_lines[cursor] == "" and new_lines[cursor] == "":
            cursor += 1
    if cursor != len(old_lines):
        return None
    return rebuilt or None


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--baseline", type=Path, required=True)
    parser.add_argument(
        "--baseline-root", type=Path, default=None,
        help="checkout whose unit data produced the baseline workbook (e.g. a "
             "worktree at c0aed960). Required once a first pass has been applied "
             "and the course rebuilt, or rows whose fix already landed stop aligning.")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--dry", action="store_true")
    args = parser.parse_args()

    reviewed = load_workbook(args.workbook, read_only=True)
    baseline = load_workbook(args.baseline, read_only=True)

    overrides: dict = {}
    if args.out.exists():
        overrides = json.loads(args.out.read_text(encoding="utf-8")).get("overrides", {})

    applied = heading_only = unmatched = unaligned = 0
    notes: list[str] = []

    for sheet in baseline.sheetnames:
        grade = int(sheet.split()[-1])
        rows_now = rows_for(grade, args.baseline_root)
        base_rows = [r for r in baseline[sheet].iter_rows(values_only=True)][1:]
        rev_rows = [r for r in reviewed[sheet].iter_rows(values_only=True)][1:]
        if len(base_rows) != len(rev_rows):
            notes.append(f"{sheet}: baseline {len(base_rows)} rows vs reviewed {len(rev_rows)} — skipped")
            continue

        for index, (was, now) in enumerate(zip(base_rows, rev_rows)):
            old_cell = str(was[4] or "")
            new_cell = str(now[4] or "")
            if old_cell == new_cell:
                continue
            item_id, category = str(was[2] or ""), str(was[1] or "")

            # The heading removal is implemented in the narration layer, so an
            # edit that is exactly that is not a content change.
            lines = old_cell.split("\n")
            if category in BOX_CATEGORIES and len(lines) > 1 \
                    and new_cell.strip() == "\n".join(lines[1:]).strip():
                heading_only += 1
                continue
            if category == GLOSSARY:
                term_prefix = old_cell.split(". ", 1)
                if len(term_prefix) == 2 and new_cell.strip() == term_prefix[1].strip():
                    heading_only += 1
                    continue

            entry = rows_now.get(item_id)
            if entry is None:
                unmatched += 1
                notes.append(f"{sheet} row {index + 2}: {category}/{item_id} has no current row — skipped")
                continue
            row, layout, source = entry

            # A box row's cell led with the title in the old export; drop it so
            # the remainder lines up with today's fields.
            adjusted = new_cell
            if category in BOX_CATEGORIES:
                title = source.get(next((p for _l, p, _k in layout if p.endswith(".title")), ""), "")
                if title and old_cell.startswith(title):
                    # The reviewer removed the title; today's layout keeps it in
                    # its own field, so only the lines are compared.
                    line_path = next((p for _l, p, _k in layout if p.endswith(".lines")), None)
                    if line_path and line_path in source:
                        if adjusted.strip() == source[line_path].strip():
                            heading_only += 1
                            continue
                        overrides.setdefault(f"grade-{grade}", {}) \
                                 .setdefault(f"unit-{row[0].split('-')[1]}", {})[line_path] = \
                            export.parse(adjusted.strip(), "lines")
                        applied += 1
                        continue

            if category == GLOSSARY:
                meaning_path = next((p for _l, p, _k in layout if p.endswith(".meaning")), None)
                if meaning_path:
                    value = adjusted.strip()
                    old_meaning = source.get(meaning_path, "")
                    if value == old_meaning:
                        heading_only += 1
                        continue
                    overrides.setdefault(f"grade-{grade}", {}) \
                             .setdefault(f"unit-{row[0].split('-')[1]}", {})[meaning_path] = value
                    applied += 1
                    continue

            manual = MANUAL_SPLITS.get(item_id)
            if manual:
                lines = new_cell.split("\n")
                edits = {}
                for path, start, stop in manual:
                    kind = next(k for _l, p, k in layout if p == path)
                    value = "\n".join(lines[start:stop]).strip()
                    if value and value != source.get(path, ""):
                        edits[path] = export.parse(value, kind)
                if edits:
                    unit_key = f"unit-{row[0].split('-')[1]}"
                    overrides.setdefault(f"grade-{grade}", {}).setdefault(unit_key, {}).update(edits)
                    applied += 1
                continue

            edits = align(old_cell, new_cell, source, layout,
                          skip_title=category in TITLE_OUTSIDE_CELL)
            if not edits:
                unaligned += 1
                notes.append(f"{sheet} row {index + 2}: {category}/{item_id} could not be aligned — skipped")
                continue
            unit_key = f"unit-{row[0].split('-')[1]}"
            for path, value in edits.items():
                kind = next(k for _l, p, k in layout if p == path)
                overrides.setdefault(f"grade-{grade}", {}).setdefault(unit_key, {})[path] = \
                    export.parse(value, kind)
            applied += 1

    payload = {
        "note": "Reviewer corrections applied over generated Global Perspectives content by "
                "tools/build-ehel-global-perspectives-runtime.js. Migrated from the "
                "pre-c0aed960 workbook format by tools/migrate-gp-legacy-review.py.",
        "source": args.workbook.name,
        "overrides": overrides,
    }
    print(f"content edits applied      : {applied}")
    print(f"heading-only (narration)   : {heading_only}   ← handled in course-ui.js, not as content")
    print(f"rows with no current match : {unmatched}")
    print(f"rows that would not align  : {unaligned}")
    for line in notes[:25]:
        print(f"  ! {line}")
    if len(notes) > 25:
        print(f"  … and {len(notes) - 25} more")
    for grade_key, units in sorted(overrides.items()):
        total = sum(len(fields) for fields in units.values())
        print(f"  {grade_key}: {total} field(s) across {len(units)} unit(s)")

    if args.dry:
        print("(dry run — nothing written)")
        return 0
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
