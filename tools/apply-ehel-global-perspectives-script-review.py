"""Turn a reviewed Global Perspectives script workbook into a builder override file.

The workbook produced by tools/export-ehel-global-perspectives-scripts.py
flattens each learner-facing item into one "Script Text" cell, joining its JSON
fields as labelled lines. A reviewer edits those cells; this script works out
which underlying field each edit belongs to and records it in

    src/prototypes/ehel-academy/global-perspectives/data/script-review.json

which tools/build-ehel-global-perspectives-runtime.js applies on top of every
rebuild. The grade data directories are GENERATED, so reviewed prose has to live
outside them or the next build would silently discard the review.

HOW IT KNOWS WHERE AN EDIT GOES
===============================
It does not keep its own map of the content. The exporter already records, per
row, the JSON path of every field in the cell (`explainers.4.body`,
`practice.11.answer`), and this reads those paths straight off that module. The
Computing equivalent maintains a parallel list of source values that has to stay
in step with its exporter row for row, guarded by an assert; here the two cannot
drift because there is only one description of the layout.

Safety, in the same spirit as the other subjects:
  * the parser is run against the UNREVIEWED text of every row and its output
    compared with the real JSON. A row whose original cannot be taken apart
    exactly is reported and skipped rather than guessed at, so a mis-split can
    never reach the content.
  * a row that drops a field entirely is reported, not applied — a reviewer
    deleting the "Answer:" line would otherwise silently erase the answer key.
  * the file is MERGED into, never replaced. Edits are found by diffing the
    workbook against the content on disk, and that content already carries any
    earlier review, so a re-run after a rebuild adds newly resolved rows instead
    of collapsing the override set to just those.

Usage:
    python tools/apply-ehel-global-perspectives-script-review.py --workbook <reviewed.xlsx>
        [--grades 1 2 ...] [--out <path.json>] [--dry]
"""

from __future__ import annotations

import argparse
import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COURSE = ROOT / "src" / "prototypes" / "ehel-academy" / "global-perspectives"
DEFAULT_OUT = COURSE / "data" / "script-review.json"

_spec = importlib.util.spec_from_file_location(
    "export_gp_scripts", ROOT / "tools" / "export-ehel-global-perspectives-scripts.py")
export = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(export)

# Categories whose text a Listen button speaks. An edit to one of these renames
# its clip, so the run reports them separately: that is the cost of the review.
NARRATED = export.NARRATED


def split_block(text: str, layout: list[tuple[str, str, str]]) -> dict[str, str] | None:
    """Take a Script Text cell apart into {json path: raw string}.

    A label occupies a whole line of its own, in brackets ("[Answer]"), and its
    value is the lines beneath it. Any field whose value was empty is absent
    entirely. An unlabelled field runs to the next label. Returns None if a
    label turns up out of order or text appears where the layout has no field to
    hold it — the caller then skips the row rather than writing a guess into the
    content.
    """
    lines = str(text or "").split("\n")

    # Locate each labelled field, in layout order, at or after the last one.
    anchors: dict[int, int] = {}
    cursor = 0
    for position, (label, _path, _kind) in enumerate(layout):
        if not label:
            continue
        marker = f"[{label}]"
        found = next((i for i in range(cursor, len(lines)) if lines[i].strip() == marker), None)
        if found is None:
            continue
        anchors[position] = found
        cursor = found + 1

    fields: dict[str, str] = {}
    line = 0
    for position, (label, path, _kind) in enumerate(layout):
        later = [anchors[p] for p in range(position + 1, len(layout)) if p in anchors]
        stop = min(later) if later else len(lines)
        if label:
            if position not in anchors:
                continue
            start = anchors[position]
            if start > line:
                return None  # unclaimed text before this label
            # The label owns its line; the value is everything beneath it.
            value = "\n".join(lines[start + 1:stop]).strip()
            line = stop
        else:
            value = "\n".join(lines[line:stop]).strip()
            line = stop
        if value:
            fields[path] = value
    if line < len(lines) and "\n".join(lines[line:]).strip():
        return None  # trailing text no field can hold
    return fields


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--grades", type=int, nargs="*")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--dry", action="store_true")
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"error: no such workbook: {args.workbook}")
        return 2

    workbook = load_workbook(args.workbook, read_only=True)
    overrides: dict = {}
    if args.out.exists():
        overrides = json.loads(args.out.read_text(encoding="utf-8")).get("overrides", {})

    changed = skipped = unsplittable = dropped = narrated_edits = 0
    notes: list[str] = []

    for sheet_name in workbook.sheetnames:
        try:
            grade = int(sheet_name.split()[-1])
        except ValueError:
            notes.append(f"{sheet_name}: not a 'Grade N' sheet — skipped")
            continue
        if args.grades and grade not in args.grades:
            continue

        files = export.unit_files(grade)
        if not files:
            notes.append(f"{sheet_name}: no unit data on disk — skipped")
            continue

        # Rebuild the exporter's rows, with the layout and source behind each.
        expected: list[tuple[str, list[str], list, dict]] = []
        for path in files:
            unit = json.loads(path.read_text(encoding="utf-8"))
            for row, layout, source in export.unit_rows_with_sources(unit):
                expected.append((path.stem, row, layout, source))

        sheet_rows = [r for r in workbook[sheet_name].iter_rows(values_only=True)][1:]
        if len(sheet_rows) != len(expected):
            notes.append(f"{sheet_name}: workbook has {len(sheet_rows)} rows, content has "
                         f"{len(expected)} — sheet skipped (re-export and re-review)")
            continue

        for index, review in enumerate(sheet_rows):
            unit_stem, row, layout, source = expected[index]
            unit_key, category, item_id, _title, original = row
            reviewed = review[4] or ""

            # The row must still be the row it was exported as. Sorting the
            # sheet, or inserting a line, would otherwise attribute an edit to
            # whatever now sits at that position.
            if str(review[0] or "") != unit_key or str(review[1] or "") != category \
                    or str(review[2] or "") != item_id:
                notes.append(f"{sheet_name} row {index + 2}: key drift "
                             f"({review[0]}/{review[1]}/{review[2]}) — skipped")
                skipped += 1
                continue

            # Prove the parser against the real JSON on every row, edited or
            # not, so a layout that has drifted shows up as a hard failure
            # rather than as silently unapplied review.
            control = split_block(original, layout)
            if control is None or control != source:
                unsplittable += 1
                notes.append(f"{sheet_name} row {index + 2}: {category}/{item_id} does not "
                             f"split cleanly — skipped")
                if reviewed != original:
                    skipped += 1
                continue
            if reviewed == original:
                continue

            parsed = split_block(reviewed, layout)
            if parsed is None:
                notes.append(f"{sheet_name} row {index + 2}: {category}/{item_id} reviewed text "
                             f"lost its shape — skipped")
                skipped += 1
                continue

            # A field present before and gone now means the reviewer deleted a
            # labelled line. Applying that would erase content — most damagingly
            # an answer key — so the row is held back whole.
            missing = [path for path in control if path not in parsed]
            if missing:
                notes.append(f"{sheet_name} row {index + 2}: {category}/{item_id} drops "
                             f"{', '.join(missing)} — not applied")
                dropped += 1
                continue

            kinds = {path: kind for _label, path, kind in layout}
            edits = {}
            for path, value in parsed.items():
                if control.get(path, "") == value:
                    continue
                edits[path] = export.parse(value, kinds[path])
            if not edits:
                continue

            overrides.setdefault(f"grade-{grade}", {}).setdefault(unit_stem, {}).update(edits)
            changed += 1
            if category in NARRATED:
                narrated_edits += 1

    payload = {
        "note": "Reviewer corrections applied over generated Global Perspectives content by "
                "tools/build-ehel-global-perspectives-runtime.js. Regenerate with "
                "tools/apply-ehel-global-perspectives-script-review.py. Keys are JSON paths "
                "into the built unit.",
        "source": args.workbook.name,
        "overrides": overrides,
    }

    print(f"rows with applied edits: {changed}   (of which narrated: {narrated_edits})")
    print(f"rows skipped: {skipped}   rows dropping a field: {dropped}")
    print(f"rows whose layout could not be verified: {unsplittable}")
    for line in notes[:40]:
        print(f"  ! {line}")
    if len(notes) > 40:
        print(f"  … and {len(notes) - 40} more")
    for grade_key, units in sorted(overrides.items()):
        total = sum(len(fields) for fields in units.values())
        print(f"  {grade_key}: {total} field(s) across {len(units)} unit(s)")
    if narrated_edits:
        print(f"\n{narrated_edits} narrated row(s) changed — those clips are now orphaned. "
              f"After rebuilding, run:\n"
              f"  node tools/generate-ehel-global-perspectives-audio.js <grade> --dry\n"
              f"  node tools/prune-ehel-course-audio.mjs global-perspectives")

    if args.dry:
        print("(dry run — nothing written)")
        return 0
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
