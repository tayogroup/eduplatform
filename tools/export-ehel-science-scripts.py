"""Export every line of learner-facing Science narration to one workbook.

Matches the layout of ehel-math-scripts-complete.xlsx so the two subjects can
be reviewed and voiced the same way:

    one sheet per grade ("Grade 1" … "Grade 8")
    columns: Unit | Category | Item ID | Title / Prompt | Script Text
    rows grouped by unit, then by category in teaching order

Science carries modules Mathematics does not (activities, fluency drills,
games, the reference panel, the "I can…" statements and a stage capstone).
Those follow the same row shape and are appended after the shared categories,
so the sheet stays complete without departing from the model.

Usage:
    python tools/export-ehel-science-scripts.py [--out <path.xlsx>]
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SCIENCE = ROOT / "src" / "prototypes" / "ehel-academy" / "science"

HEADERS = ["Unit", "Category", "Item ID", "Title / Prompt", "Script Text"]
WIDTHS = {"A": 16, "B": 20, "C": 16, "D": 30, "E": 100}
FONT = "Arial"
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
BODY_FONT = Font(name=FONT, size=10)
TOP = Alignment(vertical="top")
TOP_WRAP = Alignment(vertical="top", wrap_text=True)

# The model truncates a title derived from a prompt at 60 characters, with no
# ellipsis — the full text is in the Script Text column beside it.
TITLE_LIMIT = 60


def short(value: str) -> str:
    return str(value or "").replace("\n", " ").strip()[:TITLE_LIMIT]


def block(*pairs) -> str:
    """Join 'Label: value' lines, dropping any whose value is empty."""
    lines = []
    for label, value in pairs:
        text = value if isinstance(value, str) else ("" if value is None else str(value))
        text = text.strip()
        if not text:
            continue
        lines.append(f"{label}: {text}" if label else text)
    return "\n".join(lines)


def steps(value) -> str:
    if not value:
        return ""
    return " | ".join(str(v).strip() for v in value if str(v).strip())


def options(value) -> str:
    if not value:
        return ""
    return " / ".join(str(v).strip() for v in value if str(v).strip())


def unit_rows(u: dict) -> list[list[str]]:
    """Every narration row for one unit, in teaching order."""
    rows: list[list[str]] = []
    add = lambda category, item_id, title, script: rows.append(
        [category, str(item_id or ""), title, script])

    meta = u.get("unit", {})
    add("Unit Overview", meta.get("unitId", ""), short(meta.get("unitTitle", "")),
        block(("", meta.get("unitOverview", "")),
              ("Learning path", steps(meta.get("learningPath"))),
              ("Outcomes", steps(u.get("outcomes")))))

    for c in u.get("concepts") or []:
        add("Concept", c.get("id"), short(c.get("title")),
            block(("", c.get("explanation")), ("Example", c.get("example"))))

    for e in u.get("explorations") or []:
        add("Exploration", e.get("id"), short(e.get("title")),
            block(("", e.get("context")),
                  ("Discovery question", e.get("prompt")),
                  ("Answer", e.get("answer")),
                  ("Hint", e.get("hint")),
                  ("Explanation", e.get("explanation"))))

    for m in u.get("visualModels") or []:
        add("Visual Model", m.get("id"), short(m.get("title")), block(("", m.get("purpose"))))

    for m in u.get("methods") or []:
        add("Method", m.get("id"), short(m.get("title")),
            block(("Example", m.get("example")), ("Steps", steps(m.get("steps")))))

    for w in u.get("workedExamples") or []:
        add("Worked Example", w.get("id"), short(w.get("title")),
            block(("Prompt", w.get("prompt")), ("Solution", w.get("solution"))))

    for p in u.get("practice") or []:
        add("Practice", p.get("id"), short(p.get("prompt")),
            block(("Prompt", p.get("prompt")), ("Answer", p.get("answer")), ("Hint", p.get("hint"))))

    for i, a in enumerate(u.get("activities") or [], 1):
        add("Activity", f"activity-{i}", short(a.get("title")),
            block(("Materials", a.get("materials")), ("Steps", steps(a.get("steps")))))

    for f in u.get("fluency") or []:
        add("Fluency", f.get("id"), short(f.get("prompt")),
            block(("Prompt", f.get("prompt")), ("Answer", f.get("answer")),
                  ("Hint", f.get("hint")), ("Error feedback", f.get("errorFeedback"))))

    for r in u.get("realProblems") or []:
        add("Real Problem", r.get("id"), short(r.get("context")),
            block(("Context", r.get("context")), ("Prompt", r.get("prompt")),
                  ("Answer", r.get("answer")), ("Hint", r.get("hint")),
                  ("Error feedback", r.get("errorFeedback"))))

    for r in u.get("reasoningPrompts") or []:
        add("Reasoning Prompt", r.get("id"), short(r.get("prompt")),
            block(("Prompt", r.get("prompt")), ("Key ideas", steps(r.get("keyIdeas"))),
                  ("Model answer", r.get("modelAnswer"))))

    for q in u.get("assessment", {}).get("questions") or []:
        add("Assessment Question", q.get("id"), short(q.get("question")),
            block(("Question", q.get("question")), ("Options", options(q.get("options"))),
                  ("Answer", q.get("answer")), ("Hint", q.get("hint")),
                  ("Explanation", q.get("explanation"))))

    for g in u.get("games", {}).get("games") or []:
        for i, rnd in enumerate(g.get("rounds") or [], 1):
            add("Game Round", f"{g.get('id', 'game')}-r{i}", short(g.get("title")),
                block(("Game", g.get("title")), ("Skill", g.get("skill")),
                      ("Prompt", rnd.get("prompt")), ("Choices", options(rnd.get("choices"))),
                      ("Answer", rnd.get("answer")), ("Clue", rnd.get("clue")),
                      ("Explanation", rnd.get("explanation"))))

    ref = u.get("reference") or {}
    for i, rule in enumerate(ref.get("rules") or [], 1):
        add("Reference", f"rule-{i}", short(rule.get("title")),
            block(("Key idea", rule.get("title")), ("", rule.get("text"))))
    for i, (term, meaning) in enumerate(ref.get("terms") or [], 1):
        add("Reference", f"term-{i}", short(term), block(("Science word", term), ("Meaning", meaning)))
    for i, v in enumerate(ref.get("vocabulary") or [], 1):
        add("Reference", f"vocab-{i}", short(v.get("term")),
            block(("Word", v.get("term")), ("Meaning", v.get("meaning")), ("Example", v.get("example"))))
    for i, pair in enumerate(ref.get("commonMistakes") or [], 1):
        if len(pair) >= 2:
            add("Reference", f"mistake-{i}", short(pair[0]),
                block(("Common mistake", pair[0]), ("Correction", pair[1])))
    for i, con in enumerate(ref.get("connections") or [], 1):
        add("Reference", f"connection-{i}", short(con.get("area")),
            block(("Connects to", con.get("area")), ("", con.get("text"))))

    for i, s in enumerate(u.get("selfAssessment") or [], 1):
        add("Self Assessment", f"can-{i}", short(s), block(("", s)))

    return rows


def capstone_rows(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    cap = json.loads(path.read_text(encoding="utf-8"))
    project = cap.get("project", {})
    rows: list[list[str]] = []
    add = lambda item_id, title, script: rows.append(["Capstone", item_id, title, script])

    add("capstone-overview", short(cap.get("title")),
        block(("", cap.get("overview")),
              ("Driving question", project.get("drivingQuestion")),
              ("Final product", project.get("finalProduct"))))
    for i, st in enumerate(project.get("stages") or [], 1):
        add(f"capstone-stage-{i}", short(st.get("title")),
            block(("Stage", st.get("title")), ("", st.get("prompt"))))
    for i, item in enumerate(project.get("evidenceChecklist") or [], 1):
        add(f"capstone-evidence-{i}", short(item if isinstance(item, str) else ""),
            block(("Evidence", item if isinstance(item, str) else json.dumps(item, ensure_ascii=False))))
    for q in cap.get("quiz", {}).get("questions") or []:
        add(f"capstone-{q.get('id', '')}", short(q.get("question")),
            block(("Question", q.get("question")), ("Options", options(q.get("options"))),
                  ("Answer", q.get("answer")), ("Explanation", q.get("explanation"))))
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path,
                        default=ROOT / "outputs" / "ehel-science-scripts-complete.xlsx")
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)
    grand_total = 0
    summary = []

    for grade_dir in sorted(SCIENCE.glob("grade-*"), key=lambda p: int(p.name.split("-")[1])):
        stage = int(grade_dir.name.split("-")[1])
        unit_dir = grade_dir / "data" / "units"
        if not unit_dir.exists():
            continue
        ws = wb.create_sheet(f"Grade {stage}")
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font, cell.fill = HEAD_FONT, HEAD_FILL

        count = 0
        for path in sorted(unit_dir.glob("unit-*.json"), key=lambda p: int(p.stem.split("-")[1])):
            unit = json.loads(path.read_text(encoding="utf-8"))
            for row in unit_rows(unit):
                ws.append([path.stem, *row])
                count += 1
        for row in capstone_rows(grade_dir / "data" / "grade-capstone.json"):
            ws.append(["capstone", *row])
            count += 1

        for r in range(2, ws.max_row + 1):
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.font = BODY_FONT
                cell.alignment = TOP_WRAP if c == 5 else TOP
        for letter, width in WIDTHS.items():
            ws.column_dimensions[letter].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

        summary.append((f"Grade {stage}", count))
        grand_total += count

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    print(f"Wrote {args.out}")
    print(f"  sheets: {len(summary)}   rows: {grand_total:,}")
    for name, count in summary:
        print(f"    {name:9s} {count:6,}")


if __name__ == "__main__":
    main()
