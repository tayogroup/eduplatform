"""Turn a reviewed Computing script workbook into a builder override file.

The workbook produced by tools/export-ehel-computing-scripts.py flattens each
learner-facing item into one "Script Text" cell, joining several JSON fields
with "Label: value" lines. A reviewer edits those cells; this script works out
which underlying field each edit belongs to and records it in

    src/prototypes/ehel-academy/computing/data/script-review.json

which tools/build-ehel-computing-runtime.js applies on top of every rebuild. The
computing data directory is generated, so reviewed prose has to live outside it
or the next build would silently discard the review.

Safety: the same parser is run against the *unreviewed* text of every row and
its output is compared with the real JSON values. A row whose original text
cannot be taken apart exactly is reported and skipped rather than guessed at,
so a mis-split can never reach the content.

Usage:
    python tools/apply-ehel-computing-script-review.py --workbook <reviewed.xlsx>
        [--grades 1 2 ...] [--out <path.json>] [--dry]
"""

from __future__ import annotations

import argparse
import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COMPUTING = ROOT / "src" / "prototypes" / "ehel-academy" / "computing"
DEFAULT_OUT = COMPUTING / "data" / "script-review.json"

_spec = importlib.util.spec_from_file_location(
    "export_computing_scripts", ROOT / "tools" / "export-ehel-computing-scripts.py")
export = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(export)

# How each category's Script Text cell is assembled, in the order the exporter
# writes it: (label, field name, kind). An empty label is an unlabelled leading
# block. Kind "steps" was joined with " | ", "options" with " / ".
LAYOUTS: dict[str, list[tuple[str, str, str]]] = {
    "Unit Overview": [("", "unitOverview", "text"),
                      ("Learning path", "learningPath", "steps"),
                      ("Outcomes", "outcomes", "steps")],
    "Tool Setup": [("Tool", "name", "text"), ("Where", "url", "text"),
                   ("Steps", "steps", "steps"), ("Note", "note", "text")],
    "Concept": [("", "explanation", "text"), ("Example", "example", "text"),
                ("Check yourself", "checkYourself", "text")],
    # "Listing" restates the title and language the exporter already wrote from
    # other fields, so it has nowhere to be written back to. It is parsed so the
    # rest of the cell lines up, then dropped from the edits (see DERIVED).
    "Code Example": [("Listing", "__listing", "text"), ("", "intro", "text"),
                     ("Code (code — read for sense, do not voice verbatim)", "lines", "lines"),
                     ("Explanation", "explanation", "text")],
    "Debugging": [("Symptom", "symptom", "text"), ("Cause", "cause", "text"),
                  ("Fix", "fix", "text")],
    "Online Safety": [("Rule", "title", "text"), ("", "text", "text")],
    "AI Tutor Prompt": [("", "text", "text")],
    "Unit Project": [("Project", "title", "text"), ("", "brief", "text"),
                     ("Steps", "steps", "steps"),
                     ("Success criteria", "successCriteria", "steps")],
    "Exploration": [("", "context", "text"), ("Discovery question", "prompt", "text"),
                    ("Answer", "answer", "text"), ("Hint", "hint", "text"),
                    ("Explanation", "explanation", "text")],
    "Visual Model": [("", "purpose", "text")],
    "Method": [("Example", "example", "text"), ("Steps", "steps", "steps")],
    "Worked Example": [("Prompt", "prompt", "text"), ("Solution", "solution", "text")],
    "Practice": [("Prompt", "prompt", "text"), ("Answer", "answer", "text"),
                 ("Hint", "hint", "text")],
    "Activity": [("Materials", "materials", "text"), ("Steps", "steps", "steps")],
    "Fluency": [("Prompt", "prompt", "text"), ("Answer", "answer", "text"),
                ("Hint", "hint", "text"), ("Error feedback", "errorFeedback", "text")],
    "Real Problem": [("Context", "context", "text"), ("Prompt", "prompt", "text"),
                     ("Answer", "answer", "text"), ("Hint", "hint", "text"),
                     ("Error feedback", "errorFeedback", "text")],
    "Reasoning Prompt": [("Prompt", "prompt", "text"), ("Key ideas", "keyIdeas", "steps"),
                         ("Model answer", "modelAnswer", "text")],
    "Assessment Question": [("Question", "question", "text"), ("Options", "options", "options"),
                            ("Answer", "answer", "text"), ("Hint", "hint", "text"),
                            ("Explanation", "explanation", "text")],
    "Game Round": [("Game", "gameTitle", "text"), ("Skill", "gameSkill", "text"),
                   ("Prompt", "prompt", "text"), ("Choices", "choices", "options"),
                   ("Answer", "answer", "text"), ("Clue", "clue", "text"),
                   ("Explanation", "explanation", "text")],
    "Self Assessment": [("", "text", "text")],
    # Reference and Capstone rows vary by item id; resolved in layout_for().
    "Reference:rule": [("Key idea", "title", "text"), ("", "text", "text")],
    "Reference:term": [("Computing word", "term", "text"), ("Meaning", "meaning", "text")],
    "Reference:vocab": [("Word", "term", "text"), ("Meaning", "meaning", "text"),
                        ("Example", "example", "text")],
    "Reference:mistake": [("Common mistake", "mistake", "text"),
                          ("Correction", "correction", "text")],
    "Reference:connection": [("Connects to", "area", "text"), ("", "text", "text")],
    "Capstone:overview": [("", "overview", "text"),
                          ("Driving question", "drivingQuestion", "text"),
                          ("Final product", "finalProduct", "text")],
    "Capstone:stage": [("Stage", "title", "text"), ("", "prompt", "text"),
                       ("Evidence", "evidence", "text")],
    "Capstone:rubric": [("Criterion", "criterion", "text"), ("Secure", "secure", "text")],
    "Capstone:evidence": [("Evidence", "evidence", "text")],
    "Capstone:quiz": [("Question", "question", "text"), ("Options", "options", "options"),
                      ("Answer", "answer", "text"), ("Explanation", "explanation", "text")],
}

OPTION_FIELDS = {"options", "choices"}

# Fields the exporter renders from other values and that cannot be written
# back. An edit to one is reported, never applied.
DERIVED = {"__listing"}


def answer_holds(fields: dict[str, str]) -> bool:
    """Would this item's answer still be one of its options after the edit?

    Multiple-choice answers are stored as the full option text, so an edit that
    touches only one of the pair breaks grading — which is exactly what the
    reviewer's dash normalisation would do if the option list were held back.
    Rows that stay consistent are applied whole; rows that would not are held
    back whole and reported, so the two fields never drift apart.
    """
    answer = fields.get("answer")
    options = next((fields[f] for f in OPTION_FIELDS if f in fields), None)
    if not answer or not options:
        return True
    choices = [part.strip() for part in options.split(" / ") if part.strip()]
    return answer.strip() in choices


def layout_for(category: str, item_id: str) -> list[tuple[str, str, str]] | None:
    if category == "Reference":
        for prefix, key in (("rule-", "rule"), ("term-", "term"), ("vocab-", "vocab"),
                            ("mistake-", "mistake"), ("connection-", "connection")):
            if item_id.startswith(prefix):
                return LAYOUTS[f"Reference:{key}"]
        return None
    if category == "Capstone":
        if item_id == "capstone-overview":
            return LAYOUTS["Capstone:overview"]
        if item_id.startswith("capstone-stage-"):
            return LAYOUTS["Capstone:stage"]
        if item_id.startswith("capstone-evidence-"):
            return LAYOUTS["Capstone:evidence"]
        if item_id.startswith("capstone-rubric-"):
            return LAYOUTS["Capstone:rubric"]
        return LAYOUTS["Capstone:quiz"]
    return LAYOUTS.get(category)


def split_block(text: str, layout: list[tuple[str, str, str]]) -> dict[str, str] | None:
    """Take a Script Text cell apart into {field: raw string}.

    Labels appear in the exporter's order, each starting its own line, and any
    field whose value was empty is absent. An unlabelled field runs to the next
    label; where one directly follows a labelled field the label keeps a single
    line and the unlabelled field takes the rest, which is how the exporter's
    own rows read. Every split is checked against the real JSON before use, so a
    layout this simple rule gets wrong is skipped rather than misapplied.

    Returns None if a label turns up out of order or text appears where the
    layout has no field to hold it.
    """
    lines = str(text or "").split("\n")

    # Locate each labelled field, in layout order, at or after the last one.
    anchors: dict[int, int] = {}
    cursor = 0
    for position, (label, _field, _kind) in enumerate(layout):
        if not label:
            continue
        prefix = f"{label}: "
        found = next((i for i in range(cursor, len(lines)) if lines[i].startswith(prefix)), None)
        if found is None:
            continue
        anchors[position] = found
        cursor = found + 1

    fields: dict[str, str] = {}
    line = 0
    for position, (label, field, _kind) in enumerate(layout):
        later = [anchors[p] for p in range(position + 1, len(layout)) if p in anchors]
        stop = min(later) if later else len(lines)
        if label:
            if position not in anchors:
                continue
            start = anchors[position]
            if start > line:
                return None  # unclaimed text before this label
            # A labelled field ending an unlabelled one keeps its own line only.
            nxt = layout[position + 1] if position + 1 < len(layout) else None
            end = start + 1 if (nxt and not nxt[0] and stop > start + 1) else stop
            value = "\n".join(lines[start:end])[len(label) + 2:].strip()
            line = end
        else:
            value = "\n".join(lines[line:stop]).strip()
            line = stop
        if value:
            fields[field] = value
    if line < len(lines) and "\n".join(lines[line:]).strip():
        return None  # trailing text no field can hold
    return fields


def typed(value: str, kind: str):
    if kind == "lines":
        return value.split("\n")
    if kind == "steps":
        return [part.strip() for part in value.split(" | ") if part.strip()]
    if kind == "options":
        return [part.strip() for part in value.split(" / ") if part.strip()]
    return value


def source_fields(row_source: dict, layout: list[tuple[str, str, str]]) -> dict[str, str]:
    """The same fields as they stand in the JSON, rendered the exporter's way."""
    out: dict[str, str] = {}
    for _label, field, kind in layout:
        raw = row_source.get(field)
        if kind == "lines":
            rendered = export.listing(raw)
        elif kind == "steps":
            rendered = export.steps(raw)
        elif kind == "options":
            rendered = export.options(raw)
        else:
            rendered = str(raw or "").strip()
        if rendered:
            out[field] = rendered
    return out


def unit_sources(unit: dict) -> list[dict]:
    """One dict of the raw values behind every row of unit_rows(), same order."""
    out: list[dict] = []
    meta = unit.get("unit", {})
    out.append({"unitOverview": meta.get("unitOverview"),
                "learningPath": meta.get("learningPath"),
                "outcomes": unit.get("outcomes")})
    for t in unit.get("toolkit") or []:
        out.append({"name": t.get("name"), "url": t.get("url"),
                    "steps": t.get("steps"), "note": t.get("note")})
    for c in unit.get("concepts") or []:
        out.append({"explanation": c.get("explanation"), "example": c.get("example"),
                    "checkYourself": c.get("checkYourself")})
    for c in unit.get("codeExamples") or []:
        out.append({"__listing": f"{c.get('title')} ({c.get('language')})",
                    # Carried so the title can be split off the "Listing" line
                    # exactly, rather than guessed at from its brackets.
                    "__language": c.get("language"),
                    "intro": c.get("intro"), "lines": c.get("lines"),
                    "explanation": c.get("explanation")})
    for e in unit.get("explorations") or []:
        out.append({"context": e.get("context"), "prompt": e.get("prompt"),
                    "answer": e.get("answer"), "hint": e.get("hint"),
                    "explanation": e.get("explanation")})
    for m in unit.get("visualModels") or []:
        out.append({"purpose": m.get("purpose")})
    for m in unit.get("methods") or []:
        out.append({"example": m.get("example"), "steps": m.get("steps")})
    for w in unit.get("workedExamples") or []:
        out.append({"prompt": w.get("prompt"), "solution": w.get("solution")})
    for p in unit.get("practice") or []:
        out.append({"prompt": p.get("prompt"), "answer": p.get("answer"), "hint": p.get("hint")})
    for a in unit.get("activities") or []:
        out.append({"materials": a.get("materials"), "steps": a.get("steps")})
    for d in unit.get("debugging") or []:
        out.append({"symptom": d.get("symptom"), "cause": d.get("cause"), "fix": d.get("fix")})
    for f in unit.get("fluency") or []:
        out.append({"prompt": f.get("prompt"), "answer": f.get("answer"),
                    "hint": f.get("hint"), "errorFeedback": f.get("errorFeedback")})
    for r in unit.get("realProblems") or []:
        out.append({"context": r.get("context"), "prompt": r.get("prompt"),
                    "answer": r.get("answer"), "hint": r.get("hint"),
                    "errorFeedback": r.get("errorFeedback")})
    for s in unit.get("esafety") or []:
        out.append({"title": s.get("title"), "text": s.get("text")})
    for r in unit.get("reasoningPrompts") or []:
        out.append({"prompt": r.get("prompt"), "keyIdeas": r.get("keyIdeas"),
                    "modelAnswer": r.get("modelAnswer")})
    for p in unit.get("tutorPrompts") or []:
        out.append({"text": p})
    project = unit.get("project") or {}
    if project.get("title"):
        out.append({"title": project.get("title"), "brief": project.get("brief"),
                    "steps": project.get("steps"),
                    "successCriteria": project.get("successCriteria")})
    for q in unit.get("assessment", {}).get("questions") or []:
        out.append({"question": q.get("question"), "options": q.get("options"),
                    "answer": q.get("answer"), "hint": q.get("hint"),
                    "explanation": q.get("explanation")})
    for g in unit.get("games", {}).get("games") or []:
        for rnd in g.get("rounds") or []:
            out.append({"gameTitle": g.get("title"), "gameSkill": g.get("skill"),
                        "prompt": rnd.get("prompt"), "choices": rnd.get("choices"),
                        "answer": rnd.get("answer"), "clue": rnd.get("clue"),
                        "explanation": rnd.get("explanation")})
    ref = unit.get("reference") or {}
    for rule in ref.get("rules") or []:
        out.append({"title": rule.get("title"), "text": rule.get("text")})
    for term, meaning in ref.get("terms") or []:
        out.append({"term": term, "meaning": meaning})
    for v in ref.get("vocabulary") or []:
        out.append({"term": v.get("term"), "meaning": v.get("meaning"), "example": v.get("example")})
    for pair in ref.get("commonMistakes") or []:
        if len(pair) >= 2:
            out.append({"mistake": pair[0], "correction": pair[1]})
    for con in ref.get("connections") or []:
        out.append({"area": con.get("area"), "text": con.get("text")})
    for s in unit.get("selfAssessment") or []:
        out.append({"text": s})
    return out


def capstone_sources(path: Path) -> list[dict]:
    if not path.exists():
        return []
    cap = json.loads(path.read_text(encoding="utf-8"))
    project = cap.get("project", {})
    out = [{"overview": cap.get("overview"),
            "drivingQuestion": project.get("drivingQuestion"),
            "finalProduct": project.get("finalProduct")}]
    for st in project.get("stages") or []:
        out.append({"title": st.get("title"), "prompt": st.get("prompt"),
                    "evidence": st.get("evidence")})
    for item in project.get("evidenceChecklist") or []:
        out.append({"evidence": item if isinstance(item, str) else json.dumps(item, ensure_ascii=False)})
    for r in project.get("rubric") or []:
        out.append({"criterion": r.get("criterion"), "secure": r.get("secure")})
    for q in cap.get("quiz", {}).get("questions") or []:
        out.append({"question": q.get("question"), "options": q.get("options"),
                    "answer": q.get("answer"), "explanation": q.get("explanation")})
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--grades", type=int, nargs="*")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--dry", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(args.workbook, read_only=True)
    # Edits are found by diffing the workbook against the content on disk, and
    # that content already carries any review applied by an earlier run. So the
    # file is merged into, never replaced: re-running after a rebuild adds newly
    # resolved rows instead of collapsing the override set to just those.
    overrides: dict = {}
    if args.out.exists():
        overrides = json.loads(args.out.read_text(encoding="utf-8")).get("overrides", {})
    changed = skipped = grading = unsplittable = 0
    notes: list[str] = []

    for sheet in wb.sheetnames:
        grade = int(sheet.split()[-1])
        if args.grades and grade not in args.grades:
            continue
        grade_dir = COMPUTING / f"grade-{grade}" / "data"
        if not grade_dir.exists():
            continue

        # Rebuild the exporter's rows and the raw values behind them, in step.
        keys: list[list[str]] = []
        sources: list[dict] = []
        for path in sorted((grade_dir / "units").glob("unit-*.json"),
                           key=lambda p: int(p.stem.split("-")[1])):
            unit = json.loads(path.read_text(encoding="utf-8"))
            rows = export.unit_rows(unit)
            srcs = unit_sources(unit)
            assert len(rows) == len(srcs), f"{path.stem}: {len(rows)} rows vs {len(srcs)} sources"
            for row, src in zip(rows, srcs):
                keys.append([path.stem, row[0], row[1], row[3]])
                sources.append(src)
        cap_rows = export.capstone_rows(grade_dir / "grade-capstone.json")
        cap_srcs = capstone_sources(grade_dir / "grade-capstone.json")
        assert len(cap_rows) == len(cap_srcs)
        for row, src in zip(cap_rows, cap_srcs):
            keys.append(["capstone", row[0], row[1], row[3]])
            sources.append(src)

        sheet_rows = [r for r in wb[sheet].iter_rows(values_only=True)][1:]
        if len(sheet_rows) != len(keys):
            notes.append(f"{sheet}: workbook has {len(sheet_rows)} rows, content has {len(keys)} — skipped")
            continue

        for index, review in enumerate(sheet_rows):
            unit_key, category, item_id, original = keys[index]
            reviewed = review[4] or ""
            if str(review[0] or "") != unit_key or str(review[1] or "") != category:
                notes.append(f"{sheet} row {index + 2}: key drift ({review[0]}/{review[1]}) — skipped")
                skipped += 1
                continue

            layout = layout_for(category, item_id)
            if layout is None:
                if reviewed != original:
                    notes.append(f"{sheet} row {index + 2}: no layout for {category}/{item_id} — skipped")
                    skipped += 1
                continue

            # Prove the parser against the real JSON on every row, edited or
            # not, so a layout that has drifted shows up as a hard failure
            # rather than as silently unapplied review.
            control = split_block(original, layout)
            if control is None or control != source_fields(sources[index], layout):
                unsplittable += 1
                notes.append(f"{sheet} row {index + 2}: {category}/{item_id} does not split cleanly — skipped")
                if reviewed != original:
                    skipped += 1
                continue
            if reviewed == original:
                continue
            parsed = split_block(reviewed, layout)
            if parsed is None:
                notes.append(f"{sheet} row {index + 2}: {category}/{item_id} reviewed text lost its shape — skipped")
                skipped += 1
                continue

            kinds = {field: kind for _label, field, kind in layout}
            edits = {}
            for field, value in parsed.items():
                if control.get(field, "") == value:
                    continue
                if field == "__listing":
                    # "Listing" is written as "<title> (<language>)". The title
                    # is a real field and the card's heading, so recover it when
                    # the language part is untouched; the language itself is
                    # derived from the code and is not the reviewer's to set.
                    #
                    # Split on the stored language, not on the last " (" — two
                    # of the languages contain parentheses of their own
                    # ("Python (micro:bit)"), and splitting on the bracket cut
                    # "… listing 1 (Python" into the title.
                    suffix = f" ({sources[index].get('__language', '')})"
                    old_listing = str(control.get(field, ""))
                    if suffix != " ()" and value.endswith(suffix) and old_listing.endswith(suffix):
                        new_title = value[:-len(suffix)]
                        if new_title != old_listing[:-len(suffix)]:
                            edits["title"] = new_title
                        continue
                    notes.append(f"{sheet} row {index + 2}: {category}/{item_id} edits the "
                                 f"derived language in 'Listing' — not applied")
                    skipped += 1
                    continue
                if field in DERIVED:
                    notes.append(f"{sheet} row {index + 2}: {category}/{item_id} edits "
                                 f"'{field}', which the exporter derives — not applied")
                    skipped += 1
                    continue
                edits[field] = typed(value, kinds[field])

            # An answer must stay among its options, so the pair is applied
            # together or not at all — never one half of it. Only worth asking
            # when the edit actually touched one of them: an option list can
            # contain the separator itself ("Living / Alive"), which reads as
            # inconsistent even when the row is untouched and correct.
            if (set(edits) & ({"answer"} | OPTION_FIELDS)) and not answer_holds(parsed):
                touched = sorted(set(edits) & ({"answer"} | OPTION_FIELDS))
                notes.append(f"{sheet} row {index + 2}: {category}/{item_id} would leave the answer "
                             f"outside its options ({', '.join(touched)}) — not applied")
                grading += 1
                continue
            for field in control:
                if field not in parsed:
                    notes.append(f"{sheet} row {index + 2}: {category}/{item_id} drops '{field}' — not applied")
                    skipped += 1
            if not edits:
                continue
            slot = overrides.setdefault(f"grade-{grade}", {}).setdefault(unit_key, {}) \
                            .setdefault(category, {}).setdefault(item_id, {})
            slot.update(edits)
            changed += 1

    payload = {
        "note": "Reviewer corrections applied over generated Computing content by "
                "tools/build-ehel-computing-runtime.js. Regenerate with "
                "tools/apply-ehel-computing-script-review.py.",
        "source": args.workbook.name,
        "overrides": overrides,
    }
    print(f"rows with applied edits: {changed}")
    print(f"rows skipped: {skipped}   answer/option pairs held back: {grading}")
    print(f"rows whose layout could not be verified: {unsplittable}")
    for line in notes[:40]:
        print(f"  ! {line}")
    if len(notes) > 40:
        print(f"  … and {len(notes) - 40} more")
    for grade_key, units in sorted(overrides.items()):
        total = sum(len(items) for cats in units.values() for items in cats.values())
        print(f"  {grade_key}: {total} items across {len(units)} units")

    if args.dry:
        print("(dry run — nothing written)")
        return
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
