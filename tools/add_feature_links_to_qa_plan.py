import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


BASE_URL = "https://quraan.academy"

FEATURE_LINKS = {
    "public prospective student inquiry form": "/local/hubredirect/public_intake.php",
    "public inquiry validation": "/local/hubredirect/public_intake.php",
    "public inquiry spam/security": "/local/hubredirect/public_intake.php",
    "student intake/account creation": "/local/hubredirect/student_intake.php",
    "student intake duplicate prevention": "/local/hubredirect/student_intake.php",
    "teacher intake/account creation": "/local/hubredirect/teacher_intake.php",
    "parent-student relationship": "/local/hubredirect/student_intake.php",
    "grouping criteria capture": "/local/hubredirect/live_grouping.php",
    "matching pool creation": "/local/hubredirect/live_grouping.php",
    "class group creation": "/local/hubredirect/live_grouping.php",
    "teacher availability calendar": "/local/hubredirect/live_availability.php",
    "teacher matching recommendation": "/local/hubredirect/live_grouping.php",
    "guided live session creation": "/local/hubredirect/live_create_wizard.php",
    "recurring class series creation": "/local/hubredirect/live_series_wizard.php",
    "role-based live session visibility": "/local/hubredirect/live_sessions.php",
    "bbb meeting creation": "/local/hubredirect/live_sessions.php",
    "teacher start class": "/local/hubredirect/live_sessions.php",
    "student join class": "/local/hubredirect/live_sessions.php",
    "join window enforcement": "/local/hubredirect/live_sessions.php",
    "attendance tracking": "/local/hubredirect/live_review.php",
    "teacher post-session notes": "/local/hubredirect/live_review.php",
    "homework/action plan loop": "/local/hubredirect/live_review.php",
    "parent-visible summaries": "/local/hubredirect/live_summaries.php",
    "parent trust dashboard": "/local/hubredirect/live_parent_trust.php",
    "schedule acknowledgement/read receipts": "/local/hubredirect/live_series_schedule.php",
    "recording pull from bbb": "/local/hubredirect/live_recordings_admin.php",
    "recording review/publish": "/local/hubredirect/live_recordings_admin.php",
    "recording retention/purge readiness": "/local/hubredirect/live_recordings_admin.php",
    "notifications/reminders": "/local/hubredirect/live_ops.php",
    "parent-teacher follow-up": "/local/hubredirect/live_followups.php",
    "follow-up command center": "/local/hubredirect/live_followups.php",
    "qa checklist/admin review": "/local/hubredirect/live_quality.php",
    "teacher coaching workflow": "/local/hubredirect/live_quality.php",
    "leadership escalation": "/local/hubredirect/live_leadership.php",
    "teacher improvement plan": "/local/hubredirect/live_improvement_plans.php",
    "admin diagnostics": "/local/hubredirect/live_diagnostics.php",
    "launch readiness sql": "/local/hubredirect/live_diagnostics.php",
    "privacy/access/audit logging": "/local/hubredirect/live_diagnostics.php",
}


def normalize(value):
    return " ".join(str(value or "").strip().lower().split())


def link_for_feature(feature):
    key = normalize(feature)
    if key in FEATURE_LINKS:
        return BASE_URL + FEATURE_LINKS[key]

    if "public" in key or "inquiry" in key:
        path = "/local/hubredirect/public_intake.php"
    elif "student intake" in key:
        path = "/local/hubredirect/student_intake.php"
    elif "teacher intake" in key:
        path = "/local/hubredirect/teacher_intake.php"
    elif "availability" in key:
        path = "/local/hubredirect/live_availability.php"
    elif "group" in key or "pool" in key or "matching" in key:
        path = "/local/hubredirect/live_grouping.php"
    elif "series" in key or "acknowledgement" in key or "receipt" in key:
        path = "/local/hubredirect/live_series_schedule.php"
    elif "recording" in key:
        path = "/local/hubredirect/live_recordings_admin.php"
    elif "follow" in key:
        path = "/local/hubredirect/live_followups.php"
    elif "qa" in key or "coaching" in key:
        path = "/local/hubredirect/live_quality.php"
    elif "leadership" in key:
        path = "/local/hubredirect/live_leadership.php"
    elif "improvement" in key:
        path = "/local/hubredirect/live_improvement_plans.php"
    elif "parent trust" in key:
        path = "/local/hubredirect/live_parent_trust.php"
    elif "attendance" in key or "notes" in key or "homework" in key or "summary" in key:
        path = "/local/hubredirect/live_review.php"
    elif "bbb" in key or "join" in key or "class" in key or "session" in key:
        path = "/local/hubredirect/live_sessions.php"
    else:
        path = "/local/hubredirect/live_diagnostics.php"

    return BASE_URL + path


def find_header(ws, header_name):
    wanted = normalize(header_name)
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, ws.max_column + 1):
            if normalize(ws.cell(row, col).value) == wanted:
                return row, col
    return None, None


def add_links_to_sheet(ws):
    header_row, feature_col = find_header(ws, "Feature")
    if not header_row:
        return {"sheet": ws.title, "updated": False, "reason": "No Feature header"}

    link_header_col = None
    for col in range(1, ws.max_column + 1):
        if normalize(ws.cell(header_row, col).value) == "feature link":
            link_header_col = col
            break

    if link_header_col is None:
        link_header_col = feature_col + 1
        ws.insert_cols(link_header_col)
        ws.cell(header_row, link_header_col).value = "Feature Link"
        ws.cell(header_row, link_header_col).font = Font(bold=True)

    ws.column_dimensions[ws.cell(header_row, link_header_col).column_letter].width = 46

    populated = 0
    unmatched = []
    for row in range(header_row + 1, ws.max_row + 1):
        feature = ws.cell(row, feature_col).value
        if not feature:
            continue
        url = link_for_feature(feature)
        cell = ws.cell(row, link_header_col)
        cell.value = url
        cell.hyperlink = url
        cell.style = "Hyperlink"
        populated += 1
        if normalize(feature) not in FEATURE_LINKS:
            unmatched.append(str(feature))

    return {
        "sheet": ws.title,
        "updated": True,
        "feature_col": feature_col,
        "feature_link_col": link_header_col,
        "links_populated": populated,
        "fallback_mapped": unmatched,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if not workbook.exists():
        raise FileNotFoundError(workbook)

    backup = workbook.with_name(
        workbook.stem + ".backup-" + datetime.now().strftime("%Y%m%d-%H%M%S") + workbook.suffix
    )
    shutil.copy2(workbook, backup)

    wb = load_workbook(workbook)
    results = [add_links_to_sheet(ws) for ws in wb.worksheets]
    wb.save(workbook)

    print(json.dumps({"workbook": str(workbook), "backup": str(backup), "results": results}, indent=2))


if __name__ == "__main__":
    main()
