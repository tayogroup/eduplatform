"""Export every line of learner-facing Global Perspectives text to one workbook.

Matches the layout of ehel-english-scripts-complete.xlsx (and the Science,
Mathematics and Computing workbooks) so every subject is reviewed and voiced the
same way:

    one sheet per grade ("Grade 1" … "Grade 8")
    columns: Unit | Category | Item ID | Title / Word | Script Text
    rows grouped by unit, then by category in the order the learner meets them

WHICH ROWS COST MONEY
=====================
Only some categories carry a Listen button — the unit overview, the lesson
explainers, the callout boxes and the glossary words. A clip is named by a hash
of its exact text, so editing one of those rows renames its file and orphans the
clip already paid for. Those rows are marked "(narrated)", and the run reports
new-clip and character counts per grade, so a reviewer can see what a wording
change actually costs before making it.

Everything else is on the sheet because it is still learner-facing text that
wants an eye — the activities, the practice and its answer key, the mini-project,
the toolkit — but changing it costs nothing.

THE GROWN-UP'S GUIDE
====================
Stages 1-3 carry a section written to the parent rather than the learner, and it
is deliberately kept in the adult's voice. It is exported as
"Grown-up guide (adult voice)" so a reviewer does not "correct" it into learner
voice, which is the one edit that would be wrong here.

HOW A ROW GETS BACK INTO THE CONTENT
====================================
Every cell that joins more than one JSON field labels each one ("Question: …",
"Answer: …"). Unlabelled blocks read more naturally, but two of them in a row
cannot be taken apart again, so a reviewer's edit to such a cell could never be
applied. The labels are what make the round-trip possible.

Each row also carries, alongside the text, the JSON path of every field inside
it — `explainers.4.body`, `practice.11.answer`. apply-ehel-global-perspectives-
script-review.py reads those paths straight off this module rather than keeping
its own parallel map of the content, so the two cannot drift apart. Paths, not
item ids, are what the builder writes back, so neither tool has to reproduce the
other's id scheme.

Usage:
    python tools/export-ehel-global-perspectives-scripts.py [--out <path.xlsx>]
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
COURSE = ROOT / "src" / "prototypes" / "ehel-academy" / "global-perspectives"

HEADERS = ["Unit", "Category", "Item ID", "Title / Word", "Script Text"]
# Column C is given a width here; the English template leaves it at Excel's
# default, which clips an id like "gp-g06-u01-explain-11".
WIDTHS = {"A": 16, "B": 30, "C": 26, "D": 24, "E": 100}
FONT = "Arial"
# The English workbook's maroon header, matched so the subjects look alike on a
# reviewer's screen.
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="7A1F3D")
BODY_FONT = Font(name=FONT, size=10)
TOP = Alignment(vertical="top")
TOP_WRAP = Alignment(vertical="top", wrap_text=True)

TITLE_LIMIT = 60

# Excel reads a cell whose text opens with one of these as a formula, and
# openpyxl agrees: it writes the value into the file as <f>…</f>, which Excel
# cannot parse and reports as corrupt content on open. Global Perspectives
# quotes fill-in-the-blank prompts ("______ is a global issue") and sentence
# starters, so it hits this the same way Computing's code listings do.
FORMULA_LEADS = ("=", "+", "-", "@")

# Categories whose text a Listen button speaks. Kept in step with
# tools/lib/ehel-global-perspectives-narration.js — if a category is added
# there it belongs here too, or the sheet under-reports what an edit costs.
NARRATED = {
    "Unit overview (narrated)",
    "Lesson explainer (narrated)",
    "Big idea (narrated)",
    "Worked example (narrated)",
    "AI tutor prompt (narrated)",
    "Speaking prompt (narrated)",
    "Teacher session (narrated)",
    "Reflection box (narrated)",
    # A box nested inside an activity goes through the same runtime helper as
    # the boxes above, so it is narrated too. Leaving it out while still
    # labelling the row "(narrated)" made the sheet under-report the very cost
    # the label exists to warn about.
    "Activity box (narrated)",
    "Glossary word (narrated)",
}

# Callout list → (category label, singular name), in render order.
BOX_FIELDS = [
    ("bigIdeas", "Big idea (narrated)"),
    ("models", "Worked example (narrated)"),
    ("tutorPrompts", "AI tutor prompt (narrated)"),
    ("speakingPrompts", "Speaking prompt (narrated)"),
    ("teacherSessions", "Teacher session (narrated)"),
    ("reflectionPrompts", "Reflection box (narrated)"),
]

# A field's kind decides how it is rendered into the cell and read back out.
#   text   a single string
#   lines  a list of strings, one per line
#   table  a list of {headers, rows}, rendered pipe-separated
KIND_TEXT, KIND_LINES, KIND_TABLE = "text", "lines", "table"


def write_row(ws, values: list) -> None:
    """Append a row, forcing any formula-looking cell to be stored as text."""
    ws.append(values)
    row = ws.max_row
    for column, value in enumerate(values, 1):
        if isinstance(value, str) and value[:1] in FORMULA_LEADS:
            cell = ws.cell(row=row, column=column)
            cell.data_type = "s"
            cell.quotePrefix = True


def short(value: str) -> str:
    return str(value or "").replace("\n", " ").strip()[:TITLE_LIMIT]


# A table cell can itself contain the separator: the Stage 2 survey tables hold
# tally marks, written literally as "| | | |". Splitting those on the bare pipe
# turned a two-cell row into six and silently rewrote the table, so a pipe
# inside a cell is escaped and the split ignores escaped ones.
def cell_out(value) -> str:
    return str(value).replace("\\", "\\\\").replace("|", "\\|")


def cell_in(value: str) -> str:
    return value.strip().replace("\\|", "|").replace("\\\\", "\\")


def split_cells(line: str) -> list[str]:
    cells, current, index = [], [], 0
    while index < len(line):
        char = line[index]
        if char == "\\" and index + 1 < len(line):
            current.append(line[index:index + 2])
            index += 2
            continue
        if char == "|":
            cells.append("".join(current))
            current = []
            index += 1
            continue
        current.append(char)
        index += 1
    cells.append("".join(current))
    return [cell_in(cell) for cell in cells]


def render(value, kind: str) -> str:
    """A field's value as it appears in the Script Text cell."""
    if kind == KIND_LINES:
        return "\n".join(str(v).strip() for v in (value or []) if str(v).strip())
    if kind == KIND_TABLE:
        # Tables are separated by a blank line. 21 fields in this course hold
        # more than one, and without a separator they rendered into a single
        # block that could only ever be read back as ONE table — an edit to such
        # a row would have silently merged them. The round-trip check does not
        # catch that on its own, because it compares the rendered string rather
        # than the reconstructed value.
        blocks = []
        for table in value or []:
            lines = []
            headers = table.get("headers") or []
            if headers:
                lines.append(" | ".join(cell_out(h) for h in headers))
            for row in table.get("rows") or []:
                lines.append(" | ".join(cell_out(c) for c in row))
            if lines:
                blocks.append("\n".join(lines))
        return "\n\n".join(blocks)
    return str(value or "").strip()


def parse(text: str, kind: str):
    """The inverse of render(), for reading a reviewed cell back."""
    if kind == KIND_LINES:
        return [line.strip() for line in text.split("\n") if line.strip()]
    if kind == KIND_TABLE:
        tables = []
        for block in text.split("\n\n"):
            rows = [split_cells(line) for line in block.split("\n") if line.strip()]
            if rows:
                tables.append({"headers": rows[0], "rows": rows[1:]})
        return tables
    return text.strip()


# What a Listen button actually says, keyed by item id (unique course-wide).
# Mirrors tools/lib/ehel-global-perspectives-narration.js; the reconciliation
# below proves the two agree rather than assuming it.
SPOKEN: dict[str, str] = {}


def box_spoken(box: dict) -> str:
    """box() in course-ui.js speaks the lines, space-joined — not the title.

    The title is on screen right above the button, and the script review asked
    for the repeated heading to come out of the narration. A box with no lines
    falls back to its title so it still has something to say.
    """
    lines = [str(x) for x in (box.get("lines") or [])]
    return " ".join(lines) if lines else str(box.get("title") or "")


def spoken_text(row: list[str]) -> str:
    """The exact string the Listen button speaks for this row (cell text if none)."""
    return SPOKEN.get(row[2], row[4])


def unit_rows_with_sources(unit: dict) -> list[tuple[list[str], list[tuple[str, str, str]], dict[str, str]]]:
    """Every learner-facing line of one unit, with the JSON path behind each field.

    Returns (row, layout, source) where
        row     the five workbook columns
        layout  [(label, json path, kind)] in the order the cell joins them
        source  {json path: rendered text} for the fields that were non-empty

    The apply step re-derives the layout from here, so a change to a cell's
    shape cannot leave the two tools disagreeing about how to read it.
    """
    out: list[tuple[list[str], list[tuple[str, str, str]], dict[str, str]]] = []
    meta = unit.get("unit") or {}
    label = f"unit-{meta.get('unitNo')}"
    uid = meta.get("unitId") or ""

    def add(category: str, item_id: str, title: str, *parts, spoken: str | None = None) -> None:
        """parts: (label, json path, kind, raw value).

        `spoken` is the exact string the Listen button says, where there is one.
        It is NOT the cell text: the cell labels its fields ("Title: …",
        "Lines: …") so a reviewer's edit can be read back, and those labels are
        never read aloud. Costing the cell instead of the spoken text overstated
        the bill by 27,699 characters.
        """
        layout: list[tuple[str, str, str]] = []
        source: dict[str, str] = {}
        pieces: list[str] = []
        for part_label, path, kind, raw in parts:
            layout.append((part_label, path, kind))
            # Stripped here because the cell is stripped when it is assembled
            # and again when it is read back. Recording the unstripped value
            # made 70 rows fail their own round-trip check: a table row ending
            # in an empty column renders a trailing space that the cell drops.
            text = render(raw, kind).strip()
            if not text:
                continue
            source[path] = text
            # The label sits on a line of its own, in brackets, rather than as
            # a "Table: " prefix. One of the Stage 6 toolkit items begins
            # literally "Table: best for holding neat totals…", and with the
            # prefix form the parser read that content line as the start of the
            # Table field and split the row in the wrong place. A whole line
            # equal to "[Table]" is something the prose never produces.
            pieces.append(f"[{part_label}]\n{text}" if part_label else text)
        body = "\n".join(pieces).strip()
        if not body:
            return
        # Keyed by item id, which is unique across the whole course.
        SPOKEN[item_id] = spoken if spoken is not None else body
        out.append(([label, category, item_id, short(title), body], layout, source))

    # 1. Overview — the first thing the learner reads, and a Listen button.
    add("Unit overview (narrated)", f"{uid}-overview", meta.get("unitTitle", ""),
        ("", "unit.unitOverview", KIND_TEXT, meta.get("unitOverview")),
        spoken=render(meta.get("unitOverview"), KIND_TEXT))

    # 2. Learning outcomes and the three-stage goals.
    for i, outcome in enumerate(unit.get("outcomes") or []):
        add("Learning outcome", f"{uid}-lo{i + 1:02d}", "I can …",
            ("", f"outcomes.{i}.text", KIND_TEXT, outcome.get("text")))
    for key, heading in (("starting", "Starting"), ("developing", "Developing"), ("gettingBetter", "Getting better")):
        add("Learning goals", f"{uid}-goals-{key}", heading,
            ("", f"goals.{key}", KIND_LINES, (unit.get("goals") or {}).get(key)))

    # 3. The lesson. Body, bullets and table are separate rows because only the
    #    body is narrated — editing a bullet costs nothing, editing the body
    #    orphans a clip.
    for i, explainer in enumerate(unit.get("explainers") or []):
        eid = f"{uid}-{explainer.get('id', '')}"
        title = explainer.get("title", "")
        add("Lesson explainer (narrated)", eid, title,
            ("", f"explainers.{i}.body", KIND_TEXT, explainer.get("body")),
            spoken=render(explainer.get("body"), KIND_TEXT))
        add("Lesson bullets", f"{eid}-bullets", title,
            ("", f"explainers.{i}.bullets", KIND_LINES, explainer.get("bullets")))
        add("Lesson table", f"{eid}-table", title,
            ("", f"explainers.{i}.tables", KIND_TABLE, explainer.get("tables")))

    # 4. Callout boxes. Title and lines are spoken as one clip, so they sit in
    #    one row, in that order.
    for field, category in BOX_FIELDS:
        for i, box in enumerate(unit.get(field) or []):
            add(category, f"{uid}-{box.get('id', '')}", box.get("title", ""),
                ("Title", f"{field}.{i}.title", KIND_TEXT, box.get("title")),
                ("Lines", f"{field}.{i}.lines", KIND_LINES, box.get("lines")),
                spoken=box_spoken(box))

    # 5. Reference: toolkit, checklists, words, mistakes.
    for i, card in enumerate(unit.get("toolkit") or []):
        add("Skills toolkit", f"{uid}-{card.get('id', '')}", card.get("title", ""),
            ("Title", f"toolkit.{i}.title", KIND_TEXT, card.get("title")),
            ("Intro", f"toolkit.{i}.intro", KIND_TEXT, card.get("intro")),
            ("Items", f"toolkit.{i}.items", KIND_LINES, card.get("items")),
            ("Table", f"toolkit.{i}.tables", KIND_TABLE, card.get("tables")))
    for i, check in enumerate(unit.get("checklists") or []):
        add("Checklist", f"{uid}-{check.get('id', '')}", check.get("title", ""),
            ("Title", f"checklists.{i}.title", KIND_TEXT, check.get("title")),
            ("Intro", f"checklists.{i}.intro", KIND_TEXT, check.get("intro")),
            ("Items", f"checklists.{i}.items", KIND_LINES, check.get("items")))
    reference = unit.get("reference") or {}
    for i, word in enumerate(reference.get("vocabulary") or []):
        add("Glossary word (narrated)", f"{uid}-word{i + 1:02d}", word.get("term", ""),
            ("Word", f"reference.vocabulary.{i}.term", KIND_TEXT, word.get("term")),
            ("Meaning", f"reference.vocabulary.{i}.meaning", KIND_TEXT, word.get("meaning")),
            spoken=render(word.get("meaning"), KIND_TEXT))
    add("Common mistakes", f"{uid}-mistakes", "Common mistakes to avoid",
        ("", "reference.mistakes", KIND_LINES, reference.get("mistakes")))

    # 6. The Challenge that runs through a self-study unit.
    challenge = unit.get("challenge") or {}
    add("Challenge", f"{uid}-challenge", "My Challenge",
        ("Intro", "challenge.intro", KIND_TEXT, challenge.get("intro")),
        ("Topics", "challenge.topics", KIND_LINES, challenge.get("topics")),
        ("Checkpoint", "challenge.checkpoints", KIND_LINES, challenge.get("checkpoints")))

    # 7. Activities.
    for i, activity in enumerate(unit.get("activities") or []):
        aid = f"{uid}-{activity.get('id', '')}"
        add("Activity", aid, activity.get("label") or activity.get("title", ""),
            ("Title", f"activities.{i}.label", KIND_TEXT, activity.get("label")),
            ("Intro", f"activities.{i}.intro", KIND_TEXT, activity.get("intro")),
            ("Steps", f"activities.{i}.steps", KIND_LINES, activity.get("steps")),
            ("Table", f"activities.{i}.tables", KIND_TABLE, activity.get("tables")))
        for j, box in enumerate(activity.get("boxes") or []):
            add("Activity box (narrated)", f"{aid}-box{j + 1:02d}", box.get("title", ""),
                ("Title", f"activities.{i}.boxes.{j}.title", KIND_TEXT, box.get("title")),
                ("Lines", f"activities.{i}.boxes.{j}.lines", KIND_LINES, box.get("lines")),
                spoken=box_spoken(box))

    # 8. The Stage 1-3 mini-project.
    project = unit.get("project") or {}
    if project:
        add("Mini-project", f"{uid}-project", project.get("title", "Mini-project"),
            ("", "project.intro", KIND_TEXT, project.get("intro")))
        for i, step in enumerate(project.get("steps") or []):
            add("Mini-project step", f"{uid}-{step.get('id', '')}", step.get("title", ""),
                ("Title", f"project.steps.{i}.title", KIND_TEXT, step.get("title")),
                ("Intro", f"project.steps.{i}.intro", KIND_TEXT, step.get("intro")),
                ("Items", f"project.steps.{i}.items", KIND_LINES, step.get("items")),
                ("Table", f"project.steps.{i}.tables", KIND_TABLE, step.get("tables")))

    # 9. Practice and its answer key, together, so the reviewer sees the
    #    question beside the answer that marks it.
    for i, item in enumerate(unit.get("practice") or []):
        add("Practice + answer key", f"{uid}-{item.get('id', '')}", item.get("partTitle", ""),
            ("Question", f"practice.{i}.prompt", KIND_TEXT, item.get("prompt")),
            ("Options", f"practice.{i}.options", KIND_LINES, item.get("options")),
            ("Answer", f"practice.{i}.answer", KIND_TEXT, item.get("answer")))
    for i, question in enumerate((unit.get("assessment") or {}).get("questions") or []):
        add("Unit quiz + model answer", f"{uid}-{question.get('id', '')}", short(question.get("prompt", "")),
            ("Question", f"assessment.questions.{i}.prompt", KIND_TEXT, question.get("prompt")),
            ("Model answer", f"assessment.questions.{i}.modelAnswer", KIND_TEXT, question.get("modelAnswer")))

    # 10. Reflection and self-assessment.
    for i, item in enumerate(unit.get("reflection") or []):
        add("Reflection", f"{uid}-{item.get('id', '')}", short(item.get("prompt", "")),
            ("Prompt", f"reflection.{i}.prompt", KIND_TEXT, item.get("prompt")),
            ("One way to answer", f"reflection.{i}.modelAnswer", KIND_TEXT, item.get("modelAnswer")))
    for i, item in enumerate(unit.get("selfAssessment") or []):
        add("Self-assessment", f"{uid}-{item.get('id', '')}", "I can …",
            ("", f"selfAssessment.{i}.statement", KIND_TEXT, item.get("statement")))

    # 11. The grown-up's guide — adult voice on purpose, and last, so a reviewer
    #     working top-down meets the learner's material first.
    guide = unit.get("grownUpGuide") or {}
    add("Grown-up guide (adult voice)", f"{uid}-guide-notes", "A note for the grown-up",
        ("", "grownUpGuide.notes", KIND_LINES, guide.get("notes")))
    for i, section in enumerate(guide.get("sections") or []):
        add("Grown-up guide (adult voice)", f"{uid}-{section.get('id', '')}", section.get("title", ""),
            ("Title", f"grownUpGuide.sections.{i}.title", KIND_TEXT, section.get("title")),
            ("Body", f"grownUpGuide.sections.{i}.body", KIND_TEXT, section.get("body")),
            ("Items", f"grownUpGuide.sections.{i}.items", KIND_LINES, section.get("items")),
            ("Table", f"grownUpGuide.sections.{i}.tables", KIND_TABLE, section.get("tables")))

    return out


def unit_rows(unit: dict) -> list[list[str]]:
    """Just the workbook rows, for callers that do not need the field paths."""
    return [row for row, _layout, _source in unit_rows_with_sources(unit)]


def grades_on_disk() -> list[int]:
    return sorted(int(p.name.split("-")[1]) for p in COURSE.glob("grade-*") if p.is_dir())


def unit_files(grade: int) -> list[Path]:
    unit_dir = COURSE / f"grade-{grade}" / "data" / "units"
    if not unit_dir.is_dir():
        return []
    return sorted(unit_dir.glob("unit-*.json"), key=lambda p: int(p.stem.split("-")[1]))


def build(out_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    grand_rows = grand_narrated = grand_clips = grand_chars = 0
    summary = []
    # The generator buys one clip per DISTINCT text, so two rows carrying the
    # same sentence cost once, not twice. De-duplicating on the normalised text
    # is exactly what the generator does before hashing, so this figure matches
    # what it will spend without reimplementing cyrb53 here.
    seen_texts: set[str] = set()

    for grade in grades_on_disk():
        files = unit_files(grade)
        if not files:
            continue
        sheet = workbook.create_sheet(f"Grade {grade}")
        write_row(sheet, HEADERS)
        rows = narrated_rows = clips = chars = 0
        for file in files:
            unit = json.loads(file.read_text(encoding="utf-8"))
            for row, _layout, _source in unit_rows_with_sources(unit):
                write_row(sheet, row)
                rows += 1
                if row[1] in NARRATED:
                    narrated_rows += 1
                    normalised = " ".join(spoken_text(row).split())
                    # Under 8 characters the generator skips the clip entirely.
                    if len(normalised) >= 8 and normalised not in seen_texts:
                        seen_texts.add(normalised)
                        clips += 1
                        chars += len(normalised)

        for cell in sheet[1]:
            cell.font = HEAD_FONT
            cell.fill = HEAD_FILL
            cell.alignment = TOP
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = TOP_WRAP if cell.column_letter in ("D", "E") else TOP
        for letter, width in WIDTHS.items():
            sheet.column_dimensions[letter].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{sheet.max_row}"

        summary.append((grade, rows, narrated_rows, clips, chars))
        grand_rows += rows
        grand_narrated += narrated_rows
        grand_clips += clips
        grand_chars += chars

    if not summary:
        raise SystemExit(f"error: no unit data under {COURSE} — run: npm run build:global-perspectives")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)

    print(f"wrote {out_path}")
    print(f"{'sheet':<10}{'rows':>8}{'narrated':>10}{'new clips':>11}{'chars':>12}")
    for grade, rows, narrated_rows, clips, chars in summary:
        print(f"{'Grade ' + str(grade):<10}{rows:>8}{narrated_rows:>10}{clips:>11}{chars:>12,}")
    print(f"{'total':<10}{grand_rows:>8}{grand_narrated:>10}{grand_clips:>11}{grand_chars:>12,}")
    print("\n'(narrated)' rows are the ones a Listen button speaks: editing one renames its")
    print("clip, so a change there costs a re-generation. Every other row is free to edit.")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path.home() / "OneDrive" / "Documents" / "Elevenlabs" / "ehel-global-perspectives-scripts-complete.xlsx",
    )
    args = parser.parse_args()
    build(args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
