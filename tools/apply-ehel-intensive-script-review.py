#!/usr/bin/env python3
"""Fold a reviewed Intensive English script workbook back into the authored source.

The authored files under inputs/ehel-english-intensive-source/authored/ are the
source of truth — level-N/data/ is generated from them — so corrections land
there and the next build carries them through.

Not every difference in a returned workbook is a correction, and applying the
wrong ones damages the course. Each edit is classified first:

  content    a real edit to text the authored file owns          -> applied
  prefix     the leading "Title. " this course's exporter adds
             to a narrated grammar card; the authored explanation
             never contained it                                  -> not an edit
  spacing    whitespace-only. Irrelevant to narration, load
             bearing on screen: readings and practice items use
             aligned columns, and collapsing runs of spaces
             merges table rows into each other                   -> refused
  unmapped   anything this tool cannot tie to one authored field -> reported

Rows are matched positionally against a freshly generated export, which the
tool builds by importing the exporter itself, so the two can never drift.

  python tools/apply-ehel-intensive-script-review.py --workbook <reviewed.xlsx> [--dry]
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from importlib import import_module

exporter = import_module("export-ehel-intensive-scripts".replace("-", "_")) \
    if False else None  # module name has dashes; loaded explicitly below

ROOT = Path(__file__).resolve().parent.parent
AUTHORED = ROOT / "inputs" / "ehel-english-intensive-source" / "authored"


def load_exporter():
    """Import the dash-named exporter so row generation is provably identical."""
    import importlib.util
    path = Path(__file__).resolve().parent / "export-ehel-intensive-scripts.py"
    spec = importlib.util.spec_from_file_location("ehel_intensive_export", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def ws_norm(s: str) -> str:
    return re.sub(r"[ \t]+", " ", s or "").strip()


def parse_vocabulary(script: str):
    """Split a Vocabulary cell back into meaning, example and practice list."""
    meaning = example = None
    practice: list[str] = []
    mode = None
    for line in (script or "").split("\n"):
        if line.startswith("Meaning: "):
            meaning = line[len("Meaning: "):].strip(); mode = None
        elif line.startswith("Example: "):
            example = line[len("Example: "):].strip(); mode = None
        elif line.strip() == "Practice sentences:":
            mode = "practice"
        elif mode == "practice":
            m = re.match(r"^\s*\d+\.\s*(.+)$", line)
            if m:
                practice.append(m.group(1).strip())
    return meaning, example, practice


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--dry", action="store_true", help="report only, write nothing")
    args = ap.parse_args()

    exp = load_exporter()
    course = exp.COURSE

    # Freshly generate the rows the workbook was made from, keeping each row's
    # origin (level, unit file) so an edit can be traced back to a source file.
    generated = {}
    for level_dir in sorted(course.glob("level-*"), key=lambda p: int(p.name.split("-")[1])):
        level = int(level_dir.name.split("-")[1])
        unit_dir = level_dir / "data" / "units"
        if not unit_dir.exists():
            continue
        rows = []
        for path in sorted(unit_dir.glob("unit-*.json"), key=lambda p: int(p.stem.split("-")[1])):
            unit = json.loads(path.read_text(encoding="utf-8"))
            for row in exp.unit_rows(unit):
                rows.append((path.stem, row))
        generated[f"Level {level}"] = rows

    wb = openpyxl.load_workbook(args.workbook, read_only=True)
    counts = {"content": 0, "prefix": 0, "spacing": 0, "unmapped": 0, "identical": 0}
    to_apply = []   # (level, unitStem, category, itemId, title, newScript)
    unmapped = []

    for sheet, rows in generated.items():
        if sheet not in wb.sheetnames:
            print(f"workbook has no sheet {sheet!r}; skipping")
            continue
        got = list(wb[sheet].iter_rows(min_row=2, values_only=True))
        if len(got) != len(rows):
            print(f"{sheet}: workbook has {len(got)} rows, content has {len(rows)} — "
                  "cannot match positionally, aborting")
            sys.exit(1)

        level = int(sheet.split()[1])
        for excel, (stem, row) in zip(got, rows):
            category, item_id, title, old = row
            new = excel[4] or ""
            if (excel[1], excel[2]) != (category, item_id):
                print(f"{sheet}: row key drift at {item_id}; aborting")
                sys.exit(1)
            if new == old:
                counts["identical"] += 1
                continue
            bare = title.strip().rstrip(".")
            if bare and old.startswith(bare) and ws_norm(old[len(bare):].lstrip(". ")) == ws_norm(new):
                counts["prefix"] += 1
                continue
            if ws_norm(old) == ws_norm(new):
                counts["spacing"] += 1
                continue
            if category == "Vocabulary":
                counts["content"] += 1
                to_apply.append((level, stem, category, item_id, title, new))
            else:
                counts["unmapped"] += 1
                unmapped.append(f"{sheet} {category} {item_id}")
    wb.close()

    print("Classified the returned workbook:")
    for k in ("identical", "content", "prefix", "spacing", "unmapped"):
        print(f"  {k:9s} {counts[k]:6,}")

    # Vocabulary is the one shape this tool maps mechanically: word by word,
    # keyed on the id the exporter emitted.
    changed_files = {}
    applied = skipped = 0
    for level, stem, category, item_id, title, new in to_apply:
        unit_no = int(stem.split("-")[1])
        src = AUTHORED / f"l{level}-u{unit_no:02d}.json"
        if not src.exists():
            skipped += 1
            continue
        data = changed_files.get(src) or json.loads(src.read_text(encoding="utf-8"))
        changed_files[src] = data
        meaning, example, practice = parse_vocabulary(new)
        target = None
        for group in data.get("groups", []):
            for word in group.get("words", []):
                if word.get("w") == title or word.get("w") == title.strip():
                    target = word
                    break
            if target:
                break
        if target is None or meaning is None:
            skipped += 1
            continue
        # Only the practice list is rewritten, and only when the surrounding
        # fields are untouched — anything else is not a shape this tool proved.
        if ws_norm(target.get("meaning")) != ws_norm(meaning) or \
           ws_norm(target.get("example")) != ws_norm(example):
            skipped += 1
            continue
        if practice and practice != target.get("practice"):
            target["practice"] = practice
            applied += 1
        else:
            skipped += 1

    print(f"\nVocabulary practice lists rewritten: {applied:,}   not applied: {skipped:,}")
    if unmapped:
        print(f"\nNeeds a human ({len(unmapped)}) — no single authored field to write to:")
        for u in unmapped[:20]:
            print(f"  {u}")

    if args.dry:
        print("\n--dry: nothing written.")
        return
    for src, data in changed_files.items():
        src.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {len(changed_files)} authored file(s). Rebuild with build-intensive-units.js.")


if __name__ == "__main__":
    main()
