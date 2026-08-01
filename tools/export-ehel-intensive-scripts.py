#!/usr/bin/env python3
"""Flatten every learner-facing line of Ehel Intensive English into one workbook.

One sheet per CEFR level, modelled on ehel-english-scripts-complete.xlsx: the
same columns, the same per-category row shapes, the same Arial/maroon styling.
Reviewers work in the sheet and the corrections come back the same way the
Science and Computing reviews do.

The intensive course carries content the school English course has no equivalent
for — a unit lecture, comprehension, a quiz, an assignment, self-assessment — so
those get their own categories rather than being dropped or forced into a school
category that means something else.

Teacher-facing material is deliberately excluded: teacherNotes are marked
teacher-visibility, and rubrics and outcomes are marking scaffolding, not
anything a learner ever reads.

  python tools/export-ehel-intensive-scripts.py [--out <path>]
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
COURSE = ROOT / "src" / "prototypes" / "ehel-academy" / "intensive-english"

HEADERS = ["Unit", "Category", "Item ID", "Title / Word", "Script Text", "Comment"]
WIDTHS = {"A": 16, "B": 26, "C": 30, "D": 24, "E": 100, "F": 30}
FONT = "Arial"
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="FF7A1F3D")  # as the English workbook
BODY_FONT = Font(name=FONT, size=10)
TOP = Alignment(vertical="top")
TOP_WRAP = Alignment(vertical="top", wrap_text=True)

TITLE_LIMIT = 60
# Excel refuses a cell over 32,767 characters; the longest script today is far
# short of it, so this only ever fires if the content shape changes radically.
CELL_LIMIT = 32767
# openpyxl rejects these outright rather than escaping them.
CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean(value) -> str:
    if value is None:
        return ""
    return CONTROL_CHARS.sub("", str(value)).strip()


def short(value) -> str:
    return clean(value).replace("\n", " ")[:TITLE_LIMIT]


def block(*pairs) -> str:
    """Join 'Label: value' lines, dropping any whose value is empty."""
    lines = []
    for label, value in pairs:
        text = clean(value)
        if not text:
            continue
        lines.append(f"{label}: {text}" if label else text)
    return "\n".join(lines)


def numbered(values) -> str:
    if not values:
        return ""
    items = [clean(v) for v in values if clean(v)]
    return "\n".join(f"{i}. {v}" for i, v in enumerate(items, 1))


def unit_rows(u: dict) -> list[list[str]]:
    """Every learner-facing row for one unit, in the order a learner meets it."""
    meta = u.get("unit", {})
    uid = meta.get("unitId", "")
    rows: list[list[str]] = []

    def add(category, item_id, title, script):
        text = clean(script)
        if text:
            rows.append([category, clean(item_id), short(title), text])

    # 1. What the learner reads and hears before anything else.
    add("Unit overview", f"{uid}-overview", meta.get("unitTitle"),
        block(("", meta.get("unitOverview")), ("How to work through it", meta.get("learningPath"))))
    add("Lecture (narrated)", f"{uid}-lecture", meta.get("unitTitle"),
        (u.get("visual") or {}).get("lectureScript"))

    for r in u.get("readings", []):
        add("Reading", r.get("readingId"), r.get("title"), r.get("passageScript"))

    # Comprehension carries its own answer and explanation — this course is
    # self-teaching, so the key is learner-facing and travels with the question.
    for c in u.get("comprehension", []):
        add("Comprehension + answer key", c.get("questionId"), c.get("question"),
            block(("", c.get("question")), ("Answer", c.get("correctAnswer")),
                  ("Why", c.get("explanation"))))

    for g in u.get("grammar", []):
        add("Grammar (narrated)", g.get("grammarId"), g.get("title"),
            block(("", f"{clean(g.get('title'))}. {clean(g.get('explanation'))}".strip(". ")),
                  ("Rule and examples", g.get("ruleAndExamples")),
                  ("Worked example", g.get("workedExample")),
                  ("Common mistake", g.get("commonMistake")),
                  ("Memory tip", g.get("memoryTip"))))
        add("Grammar practice + answer key", g.get("grammarId"), g.get("title"),
            block(("", g.get("practice")), ("Check yourself", g.get("answerKey"))))

    for s in u.get("speaking", []):
        add("Speaking", s.get("speakingId"), s.get("title"),
            block(("", s.get("instructionsAndModelLines")),
                  ("Practise on your own", s.get("aiTutorPrompt"))))

    for w in u.get("writing", []):
        add("Writing", w.get("writingId"), w.get("title"),
            block(("", w.get("promptAndInstructions")), ("Model", w.get("modelText")),
                  ("Sentence starter", w.get("sentenceStarter")),
                  ("Length", w.get("expectedLength")),
                  ("Success criteria", w.get("successCriteria")),
                  ("Support", w.get("support")), ("Extension", w.get("extension"))))

    for a in u.get("activities", []):
        add("Activities", a.get("activityId"), a.get("title"),
            block(("", a.get("instructionsAndItems")), ("Answers", a.get("answerSummary")),
                  ("On your own", a.get("soloPath"))))

    for q in u.get("quizzes", []):
        add("Quiz + answer key", q.get("questionId"), q.get("question"),
            block(("", q.get("question")), ("Options", q.get("options")),
                  ("Answer", q.get("correctAnswer")), ("Why", q.get("explanation"))))

    for a in u.get("assignments", []):
        add("Assignment", a.get("assignmentId"), a.get("title"), a.get("instructions"))

    for s in u.get("selfAssessment", []):
        add("Self-assessment", s.get("selfAssessmentId"), s.get("statement"),
            s.get("statement"))

    # Vocabulary last and in one block per word, exactly as the model workbook
    # lays it out: meaning, one example, then the numbered practice sentences.
    for d in u.get("dictionaryLinks", []):
        script = block(("Meaning", d.get("childMeaning")),
                       ("Example", d.get("exampleSentence")))
        practice = numbered(d.get("practiceSentences"))
        if practice:
            # The bare "Practice sentences:" heading cannot go through block(),
            # which drops any pair with an empty value — label included.
            script = "\n".join(filter(None, [script, "Practice sentences:", practice]))
        add("Vocabulary", d.get("vocabularyId"),
            d.get("displayWord") or d.get("masterWord"), script)

    return rows


def write_cell(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    # openpyxl types any string opening with '=' as a formula, and Excel reads a
    # leading +, - or @ the same way — a script line starting with a dash is
    # enough to make the whole workbook report as corrupt. Force them to text.
    if value[:1] in ("=", "+", "-", "@"):
        cell.data_type = "s"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path,
                        default=ROOT / "outputs" / "ehel-intensive-english-scripts-complete.xlsx")
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)
    summary = []
    grand_total = 0
    oversize = []

    level_dirs = sorted(COURSE.glob("level-*"), key=lambda p: int(p.name.split("-")[1]))
    for level_dir in level_dirs:
        level = int(level_dir.name.split("-")[1])
        unit_dir = level_dir / "data" / "units"
        if not unit_dir.exists():
            continue

        ws = wb.create_sheet(f"Level {level}")
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font, cell.fill = HEAD_FONT, HEAD_FILL

        count = 0
        row_no = 1
        for path in sorted(unit_dir.glob("unit-*.json"), key=lambda p: int(p.stem.split("-")[1])):
            unit = json.loads(path.read_text(encoding="utf-8"))
            for row in unit_rows(unit):
                row_no += 1
                for col, value in enumerate([path.stem, *row], start=1):
                    if len(value) > CELL_LIMIT:
                        oversize.append(f"{path.stem} {row[1]} ({len(value):,} chars)")
                        value = value[:CELL_LIMIT]
                    write_cell(ws, row_no, col, value)
                count += 1

        for r in range(2, ws.max_row + 1):
            for c in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = BODY_FONT
                cell.alignment = TOP_WRAP if c == 5 else TOP
        for letter, width in WIDTHS.items():
            ws.column_dimensions[letter].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

        summary.append((f"Level {level}", count))
        grand_total += count

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    print(f"Wrote {args.out}")
    print(f"  sheets: {len(summary)}   rows: {grand_total:,}")
    for name, count in summary:
        print(f"    {name:9s} {count:6,}")
    if oversize:
        print(f"  TRUNCATED at Excel's {CELL_LIMIT:,}-character cell limit:")
        for item in oversize:
            print(f"    {item}")


if __name__ == "__main__":
    main()
