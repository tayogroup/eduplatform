from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("outputs/mock_student_grouping")
CSV_PATH = OUTPUT_DIR / "quraan_academy_50_student_grouping_mock_data.csv"
XLSX_PATH = OUTPUT_DIR / "quraan_academy_50_student_grouping_mock_data.xlsx"


HEADERS = [
    "mock_student_no",
    "student_display_name",
    "student_firstname",
    "student_lastname",
    "age",
    "over_18",
    "gender",
    "special_needs",
    "course_type",
    "country",
    "city",
    "timezone",
    "timezone_group",
    "primary_language",
    "other_languages",
    "current_level",
    "base_of_learning",
    "number_of_sessions_per_week",
    "preferred_days",
    "preferred_hours_local",
    "schedule_choices",
    "parent_guardian_required",
    "parent_name",
    "parent_contact_email_or_phone",
    "parent_phone_whatsapp",
    "live_class_consent",
    "recording_consent",
    "consent_notes_comment",
    "recommended_group_pool",
    "recommended_group_size_target",
    "bbb_ready",
    "moodle_course_group_ready",
    "admin_notes",
]


TIMEZONE_SEGMENTS = [
    {
        "count": 11,
        "country": "United States",
        "city_pool": ["Minneapolis", "Saint Paul", "Rochester", "Bloomington", "Duluth"],
        "timezone": "America/Chicago",
        "timezone_group": "Minnesota",
        "primary_language": "English",
        "other_languages": ["Somali", "Arabic"],
        "hours": ["5:00 PM", "6:00 PM", "7:00 PM", "8:00 PM"],
    },
    {
        "count": 13,
        "country": "Kenya",
        "city_pool": ["Nairobi", "Mombasa", "Kisumu", "Eldoret", "Garissa"],
        "timezone": "Africa/Nairobi",
        "timezone_group": "Kenya",
        "primary_language": "Somali",
        "other_languages": ["English", "Arabic", "Swahili"],
        "hours": ["4:00 PM", "5:00 PM", "6:00 PM", "7:00 PM", "8:00 PM"],
    },
    {
        "count": 10,
        "country": "United Kingdom",
        "city_pool": ["London", "Birmingham", "Manchester", "Leicester", "Bristol"],
        "timezone": "Europe/London",
        "timezone_group": "London",
        "primary_language": "English",
        "other_languages": ["Somali", "Arabic"],
        "hours": ["5:00 PM", "6:00 PM", "7:00 PM", "8:00 PM"],
    },
    {
        "count": 9,
        "country": "Pakistan",
        "city_pool": ["Lahore", "Karachi", "Islamabad", "Faisalabad", "Rawalpindi"],
        "timezone": "Asia/Karachi",
        "timezone_group": "Lahore, Pakistan",
        "primary_language": "Urdu",
        "other_languages": ["English", "Arabic"],
        "hours": ["6:00 PM", "7:00 PM", "8:00 PM", "9:00 PM", "10:00 PM"],
    },
    {
        "count": 7,
        "country": "Saudi Arabia",
        "city_pool": ["Riyadh", "Jeddah", "Makkah", "Madinah", "Dammam"],
        "timezone": "Asia/Riyadh",
        "timezone_group": "Saudi Arabia",
        "primary_language": "Arabic",
        "other_languages": ["English", "Somali"],
        "hours": ["5:00 PM", "6:00 PM", "7:00 PM", "8:00 PM", "9:00 PM"],
    },
]


FIRST_NAMES = [
    "Ayaan", "Maryam", "Yusuf", "Fatima", "Omar", "Hana", "Ibrahim", "Zahra", "Adam", "Safiya",
    "Musa", "Amina", "Bilal", "Layla", "Hamza", "Nura", "Ismail", "Sumaya", "Khalid", "Hodan",
    "Salman", "Asma", "Yasir", "Ruqiya", "Idris", "Amal", "Anas", "Sahra", "Zayd", "Naima",
    "Hassan", "Khadija", "Ali", "Muna", "Abdi", "Rahma", "Tariq", "Ilhan", "Kareem", "Faduma",
    "Mahad", "Zainab", "Abdullah", "Iman", "Farhan", "Halima", "Sami", "Nasra", "Aisha", "Rayyan",
]

LAST_NAMES = [
    "Ahmed", "Yusuf", "Hassan", "Ali", "Mohamed", "Osman", "Abdi", "Nur", "Ibrahim", "Farah",
]


def pick(items: list[str], index: int) -> str:
    return items[index % len(items)]


def build_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    number = 1
    day_patterns = [
        ["Monday", "Wednesday"],
        ["Tuesday", "Thursday"],
        ["Saturday"],
        ["Sunday", "Wednesday"],
        ["Monday", "Friday"],
    ]
    levels = ["alphabet", "level_1", "level_2", "short_surahs", "tajweed_intro"]
    bases = ["new_learner", "knows_letters", "can_read_simple_words", "needs_revision", "memorization_beginner"]
    courses = ["Pre-quraan Course", "Quraan Memorization Course"]

    for segment in TIMEZONE_SEGMENTS:
        for local_index in range(segment["count"]):
            first = pick(FIRST_NAMES, number - 1)
            last = pick(LAST_NAMES, number + local_index)
            age = 6 + ((number + local_index) % 13)
            if number in {12, 31}:
                age = 19
            over_18 = "yes" if age >= 18 else "no"
            gender = "female" if number % 2 == 0 else "male"
            sessions = 1 + (number % 5)
            days = pick(day_patterns, number)
            hours = [pick(segment["hours"], number + offset) for offset in range(min(sessions, len(days) + 1))]
            schedule_choices = "; ".join(f"{day} {pick(hours, idx)}" for idx, day in enumerate(days))
            parent_required = "no" if over_18 == "yes" else "yes"
            parent_name = "" if parent_required == "no" else f"{pick(['Abdirahman', 'Khadra', 'Hussein', 'Zamzam', 'Omar'], number)} {last}"
            parent_contact = (
                f"+1-555-01{number:03d}"
                if parent_required == "no"
                else f"parent{number:02d}@example.test"
            )

            current_level = pick(levels, number + local_index)
            base = pick(bases, number)
            course = courses[number % 2]
            group_pool = (
                f"{segment['timezone_group']} | {course} | {current_level} | "
                f"{gender} | ages {max(6, age - 1)}-{min(18, age + 1)}"
            )

            rows.append({
                "mock_student_no": number,
                "student_display_name": f"{first} {last}",
                "student_firstname": first,
                "student_lastname": last,
                "age": age,
                "over_18": over_18,
                "gender": gender,
                "special_needs": "yes" if number in {7, 24, 41} else "no",
                "course_type": course,
                "country": segment["country"],
                "city": pick(segment["city_pool"], local_index),
                "timezone": segment["timezone"],
                "timezone_group": segment["timezone_group"],
                "primary_language": segment["primary_language"],
                "other_languages": ", ".join(segment["other_languages"]),
                "current_level": current_level,
                "base_of_learning": base,
                "number_of_sessions_per_week": sessions,
                "preferred_days": ", ".join(days),
                "preferred_hours_local": ", ".join(hours),
                "schedule_choices": schedule_choices,
                "parent_guardian_required": parent_required,
                "parent_name": parent_name,
                "parent_contact_email_or_phone": parent_contact,
                "parent_phone_whatsapp": f"+{10000000000 + number}",
                "live_class_consent": "yes",
                "recording_consent": "yes" if number % 4 != 0 else "no",
                "consent_notes_comment": "Mock consent captured for planning only.",
                "recommended_group_pool": group_pool,
                "recommended_group_size_target": 9,
                "bbb_ready": "yes",
                "moodle_course_group_ready": "yes",
                "admin_notes": "Mock record for grouping, scheduling, and BBB/Moodle alignment.",
            })
            number += 1

    return rows


def write_csv(rows: list[dict[str, object]]) -> None:
    with CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Mock Students"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row[header] for header in HEADERS])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 12, "B": 22, "C": 16, "D": 16, "E": 8, "F": 10, "G": 10, "H": 13,
        "I": 28, "J": 18, "K": 18, "L": 20, "M": 18, "N": 18, "O": 24, "P": 18,
        "Q": 24, "R": 16, "S": 24, "T": 24, "U": 34, "V": 22, "W": 22, "X": 28,
        "Y": 20, "Z": 16, "AA": 18, "AB": 32, "AC": 48, "AD": 18, "AE": 12,
        "AF": 22, "AG": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = workbook.create_sheet("Summary")
    summary_rows = [
        ["Timezone Group", "Student Count"],
        ["Minnesota", 11],
        ["Kenya", 13],
        ["London", 10],
        ["Lahore, Pakistan", 9],
        ["Saudi Arabia", 7],
        ["Total", 50],
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 16

    workbook.save(XLSX_PATH)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = build_rows()
    if len(rows) != 50:
        raise RuntimeError(f"Expected 50 rows, got {len(rows)}")
    counts = {}
    for row in rows:
        counts[row["timezone_group"]] = counts.get(row["timezone_group"], 0) + 1
    expected = {
        "Minnesota": 11,
        "Kenya": 13,
        "London": 10,
        "Lahore, Pakistan": 9,
        "Saudi Arabia": 7,
    }
    if counts != expected:
        raise RuntimeError(f"Unexpected timezone counts: {counts}")
    write_csv(rows)
    write_xlsx(rows)
    print(f"Wrote {CSV_PATH}")
    print(f"Wrote {XLSX_PATH}")


if __name__ == "__main__":
    main()
